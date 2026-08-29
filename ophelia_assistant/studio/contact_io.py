"""Contact file import/export helpers for the studio."""

from __future__ import annotations

import csv
from pathlib import Path


def _index_of(headers: list[str], *names: str) -> int:
    lowered = [str(header or "").strip().lower() for header in headers]
    for name in names:
        if name.lower() in lowered:
            return lowered.index(name.lower())
    return -1


def _cell(row, index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    return str(row[index] or "").strip()


def _normalize_entries(headers: list[str], raw_rows) -> list[dict[str, str]]:
    name_idx = _index_of(headers, "名字", "name", "姓名")
    location_idx = _index_of(headers, "地区", "location", "城市", "city")
    email_idx = _index_of(headers, "邮箱", "email", "邮箱地址", "mail")
    if name_idx < 0 or location_idx < 0 or email_idx < 0:
        raise ValueError("文件缺少必要表头：名字、地区、邮箱")
    entries: list[dict[str, str]] = []
    for row in raw_rows:
        name = _cell(row, name_idx)
        location = _cell(row, location_idx)
        email = _cell(row, email_idx)
        if not (name or location or email):
            continue
        if (
            name in {"张三", "John"}
            and location == "Seattle"
            and email in {"zhangsan@example.com", "john@example.com"}
        ):
            continue
        if name.startswith("示例") or name.startswith("模板"):
            continue
        entry = {"name": name, "location": location, "email": email}
        for index, header in enumerate(headers):
            key = str(header or "").strip().lower()
            if key.startswith("custom_") or key.startswith("变量"):
                value = _cell(row, index)
                if value:
                    entry[key] = value
        entries.append(entry)
    return entries


def parse_contacts_file(path: str) -> list[dict[str, str]]:
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    if suffix == ".csv":
        return _parse_csv(path)
    raise ValueError("仅支持 .xlsx 或 .csv 文件")


def _parse_xlsx(path: str) -> list[dict[str, str]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    if not raw_rows:
        return []
    headers = [str(value or "").strip() for value in raw_rows[0]]
    return _normalize_entries(headers, raw_rows[1:])


def _parse_csv(path: str) -> list[dict[str, str]]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = [list(row) for row in reader]
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return _normalize_entries(headers, rows[1:])


def write_import_template(path: str) -> None:
    """Write the standard contact import workbook with headers and one example."""
    import openpyxl
    from openpyxl.styles import Font

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "联系人"
    headers = ["名字", "地区", "邮箱"] + [
        f"custom_{index}" for index in range(1, 6)
    ]
    sheet.append(headers)
    for column in sheet[1]:
        column.font = Font(bold=True)
    sheet.append(
        ["张三", "Seattle", "zhangsan@example.com", "", "", "", "", ""]
    )
    for column, width in zip("ABCDEFGH", (16, 16, 28, 14, 14, 14, 14, 14)):
        sheet.column_dimensions[column].width = width
    workbook.save(path)

