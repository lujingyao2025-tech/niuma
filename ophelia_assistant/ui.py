# THESIS: 牛马邮箱是一张可调度、可追踪、可恢复的发信任务台，拒绝传统后台把功能平铺成菜单与表格。
# OWN-WORLD: 深海军蓝导航、冷白工作面、信号蓝主动作、绿色完成态与琥珀提醒；组件像签派清单与状态灯。
# STORY: 用户按导入、配置、执行、完成四段轨迹推进一批联系人，并随时看清任务数量、风险与下一步。
# FIRST VIEWPORT: 224px 导航固定在左；顶部任务轨迹横贯主区；联系人工作台居中；关键发送动作保持高辨识度。
# FORM: 任务调度台，候选方向第六项；seed 28f129ba。签名交互是随任务状态点亮的四段发送轨迹。
# FINISH: unreviewed and undocumented is unfinished; this build ends with the finish review, the verdict, DESIGN.md, and every shipping raster carrying its provenance
from __future__ import annotations

import json
import logging
import os
import re
import sys
import tempfile
import threading
import time
import tkinter as tk
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from tkinter import colorchooser, filedialog, messagebox, ttk

import requests

from PIL import Image, ImageOps, ImageTk

from . import __version__
from .config import (
    BATCH_CONTACT_ROWS,
    DEFAULT_BODY_TEMPLATE,
    DEFAULT_SENDER_NAME,
    DEFAULT_SUBJECT_TEMPLATE,
    MAX_CONCURRENT_TASKS,
    MAX_CONTACT_ROWS,
    MAX_WINDOW_SEQUENCE,
    Settings,
    is_newer_version,
    next_custom_variable_key,
    normalize_window_sequence,
    resolve_task_windows,
    resolve_task_windows_balanced,
)
from .database import Database, now_iso
from .i18n import LANGUAGES, current_language, set_language, tr, trf
from .morelogin import BROWSER_PROVIDER_NAMES, create_browser_provider
from .operation import OperationCancelledError, OperationController
from .trial import check_trial, device_code, remaining_text, verify_authorization_code
from .workflow import Workflow


logger = logging.getLogger("niuma-mail")


STATUS_LABELS = {
    "new": "新任务",
    "ready": "待确认",
    "pending": "待处理",
    "generated": "已生成",
    "filling": "正在填写",
    "needs_review": "需要手动修改",
    "drafted": "Gmail 草稿",
    "sending": "正在发送",
    "sent": "历史：已发送",
    "replied": "历史：已回复",
    "failed": "发送失败",
    "cancelled": "已取消",
}
SYSTEM_VARIABLES = (
    ("first_name", "联系人称呼（完整姓名时使用姓氏）"),
    ("location", "城市或城市地区"),
    ("sender_name", "发件人姓名"),
)
INK = "#111827"
MUTED = "#667085"
PAPER = "#FFFFFF"
PANEL = "#F3F6FA"
LINE = "#DCE3EC"
GOLD = "#2F6BFF"
GOLD_DARK = "#1E4FD6"
RED = "#EF4444"
GREEN = "#10B981"
HEADER = "#0B1220"
SIDEBAR = "#0B1220"
SIDEBAR_HOVER = "#17243A"
SURFACE = "#F8FAFC"
AMBER = "#F59E0B"
SKY = "#0EA5E9"
APP_VERSION = f"v{__version__}"
DISPLAY_FONT = "Bahnschrift SemiCondensed"


DEFAULT_THEMES: dict[str, dict[str, str]] = {
    "light": {
        "ink": "#111827",
        "muted": "#667085",
        "paper": "#FFFFFF",
        "panel": "#F3F6FA",
        "line": "#DCE3EC",
        "gold": "#2F6BFF",
        "gold_dark": "#1E4FD6",
        "red": "#EF4444",
        "green": "#16A34A",
        "header": "#0B1220",
        "sidebar": "#0B1220",
        "sidebar_hover": "#17243A",
        "surface": "#F8FAFC",
        "amber": "#F59E0B",
        "sky": "#0EA5E9",
    },
    "dark": {
        "ink": "#E8EEF8",
        "muted": "#8FA1BC",
        "paper": "#141F36",
        "panel": "#0D1526",
        "line": "#26354F",
        "gold": "#4D8DFF",
        "gold_dark": "#2F6BE0",
        "red": "#F87171",
        "green": "#34D399",
        "header": "#0A1122",
        "sidebar": "#0A1122",
        "sidebar_hover": "#12203A",
        "surface": "#18243D",
        "amber": "#FBBF24",
        "sky": "#38BDF8",
    },
}


def apply_theme(settings) -> None:
    """Apply the selected theme and custom skin to the UI color tokens."""
    global INK, MUTED, PAPER, PANEL, LINE, GOLD, GOLD_DARK
    global RED, GREEN, HEADER, SIDEBAR, SIDEBAR_HOVER, SURFACE, AMBER, SKY
    tokens = dict(DEFAULT_THEMES.get(settings.theme_mode, DEFAULT_THEMES["light"]))
    tokens.update(settings.skin_colors or {})
    INK = tokens["ink"]
    MUTED = tokens["muted"]
    PAPER = tokens["paper"]
    PANEL = tokens["panel"]
    LINE = tokens["line"]
    GOLD = tokens["gold"]
    GOLD_DARK = tokens["gold_dark"]
    RED = tokens["red"]
    GREEN = tokens["green"]
    HEADER = tokens["header"]
    SIDEBAR = tokens["sidebar"]
    SIDEBAR_HOVER = tokens["sidebar_hover"]
    SURFACE = tokens["surface"]
    AMBER = tokens["amber"]
    SKY = tokens["sky"]


class _ReusableExecutor:
    """A shared pool that existing coordinator code can use as a context."""

    def __init__(self, max_workers: int) -> None:
        self.executor = ThreadPoolExecutor(max_workers=max_workers)

    def __enter__(self):
        return self.executor

    def __exit__(self, _exc_type, _exc, _traceback) -> bool:
        return False

    def submit(self, *args, **kwargs):
        return self.executor.submit(*args, **kwargs)

    def shutdown(self) -> None:
        self.executor.shutdown(wait=False, cancel_futures=True)


def parse_contacts_xlsx(path: str) -> list[dict[str, str]]:
    """Read a contact XLSX into normalized dict rows, skipping the template row."""
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    if not raw_rows:
        return []
    headers = [str(value or "").strip() for value in raw_rows[0]]
    lowered_headers = [str(value).lower() for value in headers]

    def index_of(*names: str) -> int:
        for name in names:
            if name.lower() in lowered_headers:
                return lowered_headers.index(name.lower())
        return -1

    name_idx = index_of("名字", "name", "姓名")
    location_idx = index_of("地区", "location", "城市", "city")
    email_idx = index_of("邮箱", "email", "邮箱地址", "mail")
    if name_idx < 0 or location_idx < 0 or email_idx < 0:
        raise ValueError("XLSX 模板缺少必要表头：名字、地区、邮箱")

    def cell(row: list, index: int) -> str:
        if index < 0 or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    entries: list[dict[str, str]] = []
    for row in raw_rows[1:]:
        name = cell(row, name_idx)
        location = cell(row, location_idx)
        email = cell(row, email_idx)
        if not (name or location or email):
            continue
        # The bundled template's example row is display-only and never imported.
        if (
            name == "张三"
            and location == "Seattle"
            and email == "zhangsan@example.com"
        ) or (
            name == "John"
            and location == "Seattle"
            and email == "john@example.com"
        ):
            continue
        if name.startswith("示例") or name.startswith("模板"):
            continue
        entry = {"name": name, "location": location, "email": email}
        for index, header in enumerate(headers):
            key = header.strip().lower()
            if key.startswith("custom_") or key.startswith("变量"):
                entry[key] = cell(row, index)
        entries.append(entry)
    return entries


def resource_path(relative: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent.parent))
    return base / relative


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"{tr('牛马邮箱')} · {APP_VERSION}")
        self.geometry("1360x860")
        self.minsize(1120, 720)
        self.configure(bg=PANEL)

        self._resize_job = None
        self._refresh_job = None
        self._trial_job = None
        self._row_snapshots: dict[int, tuple] = {}
        self._all_rows = []
        self._load_brand_assets()
        self.settings = Settings.load()
        set_language(self.settings.language)
        apply_theme(self.settings)
        self._configure_styles()
        self._build_background()
        self._trial_status = check_trial()
        self.db = Database()
        self.db.backup()
        self._last_failed_task_ids: set[int] = set()
        self._patch_messagebox()
        self.workflow = Workflow(self.db, self.settings)
        self.operations = OperationController()
        self._operation_busy = False
        self._operation_buttons: list[ttk.Button] = []
        self._test_busy = False
        self._busy_watchdog_job = None
        self._worker_pool = _ReusableExecutor(MAX_CONCURRENT_TASKS)
        self._build_workspace()
        self._set_operation_busy(False)
        self.after_idle(self.refresh)
        self._translate_widgets()
        self._apply_authorization_state()
        if not self._trial_status.active:
            self.after_idle(self._show_authorization_required)
        else:
            self.after_idle(self._prompt_unfinished_tasks)
        self.after_idle(lambda: self.check_for_update(silent=True))
        self._trial_job = self.after(60_000, self._enforce_trial)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _translate_widgets(self) -> None:
        """Apply the current UI language to static widgets without rebuilding."""
        from .i18n import current_language, tr

        english = current_language() == "en"
        cjk = re.compile(r"[\u4e00-\u9fff]")

        def walk(widget) -> None:
            try:
                if isinstance(widget, (tk.Label, tk.Button, ttk.Label, ttk.Button)):
                    original = getattr(widget, "_niuma_orig_text", None)
                    if original is None:
                        original = str(widget.cget("text") or "")
                        widget._niuma_orig_text = original
                    text = tr(original) if english else original
                    if str(widget.cget("text") or "") != text:
                        widget.configure(text=text)
            except tk.TclError:
                pass
            try:
                children = widget.winfo_children()
            except tk.TclError:
                children = []
            for child in children:
                walk(child)

        walk(self)

        if not hasattr(self, "_tab_text_orig"):
            self._tab_text_orig = {}
        for notebook in (
            getattr(self, "notebook", None),
            getattr(self, "settings_sections", None),
        ):
            if notebook is None:
                continue
            try:
                tab_ids = notebook.tabs()
            except tk.TclError:
                continue
            for tab_id in tab_ids:
                original = self._tab_text_orig.get(tab_id)
                if original is None:
                    original = str(notebook.tab(tab_id, "text") or "")
                    self._tab_text_orig[tab_id] = original
                notebook.tab(tab_id, text=tr(original) if english else original)

        if not hasattr(self, "_heading_text_orig"):
            self._heading_text_orig = {}
        for tree in (
            getattr(self, "tree", None),
            getattr(self, "history_tree", None),
        ):
            if tree is None:
                continue
            tree_id = str(tree)
            try:
                columns = list(tree["columns"])
            except tk.TclError:
                continue
            for column in columns:
                original = self._heading_text_orig.get((tree_id, column))
                if original is None:
                    original = str(tree.heading(column, "text") or "")
                    self._heading_text_orig[(tree_id, column)] = original
                tree.heading(column, text=tr(original) if english else original)

        for attr in (
            "status_var",
            "connection_var",
            "selection_count_var",
            "profile_assign_note_var",
            "window_sequence_note_var",
            "ungenerated_hint_var",
        ):
            variable = getattr(self, attr, None)
            if variable is not None:
                value = str(variable.get() or "")
                if english:
                    original = getattr(variable, "_niuma_orig_value", None)
                    if original is None:
                        original = value
                        variable._niuma_orig_value = original
                    if cjk.search(original):
                        variable.set(tr(original))
                else:
                    original = getattr(variable, "_niuma_orig_value", None)
                    if original is not None and value != original:
                        variable.set(original)

    def _patch_messagebox(self) -> None:
        import tkinter.messagebox as mb

        for name in (
            "showinfo",
            "showwarning",
            "showerror",
            "askyesno",
            "askyesnocancel",
        ):
            original = getattr(mb, name)
            if getattr(original, "_niuma_i18n_patched", False):
                continue
            setattr(mb, f"_orig_{name}", original)

            def make_wrapper(kind: str):
                def wrapped(*args, **kwargs):
                    args = tuple(tr(str(arg)) for arg in args)
                    for key in ("title", "message"):
                        if key in kwargs:
                            kwargs[key] = tr(str(kwargs[key]))
                    return getattr(mb, f"_orig_{kind}")(*args, **kwargs)

                return wrapped

            wrapped = make_wrapper(name)
            wrapped._niuma_i18n_patched = True
            setattr(mb, name, wrapped)

    def _prompt_unfinished_tasks(self) -> None:
        try:
            rows = self.db.list_tasks()
        except Exception:
            return
        active = [row for row in rows if row["status"] not in {"sent", "replied"}]
        if not active:
            return
        ungenerated = sum(
            1 for row in active if row["status"] in {"pending", "new", "needs_review"}
        )
        drafted = sum(1 for row in active if row["status"] == "drafted")
        ready = sum(
            1 for row in active if row["status"] in {"generated", "ready"}
        )
        if not messagebox.askyesno(
            "继续上次任务",
            f"检测到 {len(active)} 条未完成任务：\n"
            f"未生成 {ungenerated} 条、待确认 {ready} 条、"
            f"Gmail 草稿 {drafted} 条。\n\n"
            "是否选中这些任务，方便继续处理？",
            default=messagebox.YES,
        ):
            return
        self._select_task_ids([int(row["id"]) for row in active])
        self._select_page(self.queue_tab)
        self._set_status(f"已选中 {len(active)} 条未完成任务，可继续生成或填写")

    def _load_brand_assets(self) -> None:
        try:
            self.iconbitmap(str(resource_path("assets/niuma-mail-icon.ico")))
            icon = Image.open(resource_path("assets/niuma-mail-icon.png")).convert("RGBA")
            self._window_icon = ImageTk.PhotoImage(icon.resize((64, 64), Image.Resampling.LANCZOS))
            self.iconphoto(True, self._window_icon)
            self._header_icon = self._window_icon
        except (OSError, tk.TclError):
            self._window_icon = None
            self._header_icon = None
        self._background_source = None
        try:
            contact_qr = Image.open(resource_path("assets/telegram-ls0514.png")).convert("RGB")
            contact_qr = ImageOps.contain(contact_qr, (230, 290), method=Image.Resampling.LANCZOS)
            self._contact_qr_photo = ImageTk.PhotoImage(contact_qr)
        except (OSError, tk.TclError):
            self._contact_qr_photo = None

    def _configure_styles(self) -> None:
        style = ttk.Style(self)
        style.theme_use("clam")
        style.layout("Sidebar.TNotebook.Tab", [])
        style.configure("Sidebar.TNotebook", background=PANEL, borderwidth=0, tabmargins=0)
        style.configure("App.TNotebook", background=PANEL, borderwidth=0, tabmargins=(12, 10, 12, 0))
        style.configure(
            "App.TNotebook.Tab",
            background="#E9EEF5",
            foreground=MUTED,
            padding=(22, 11),
            font=("Microsoft YaHei UI", 10, "bold"),
            borderwidth=0,
        )
        style.map("App.TNotebook.Tab", background=[("selected", PAPER)], foreground=[("selected", INK)])
        style.configure(
            "App.Treeview",
            background=PAPER,
            fieldbackground=PAPER,
            foreground=INK,
            rowheight=42,
            borderwidth=0,
            font=("Microsoft YaHei UI", 9),
        )
        style.configure(
            "App.Treeview.Heading",
            background="#EDF1F6",
            foreground=INK,
            relief="flat",
            padding=(10, 10),
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        style.map("App.Treeview", background=[("selected", "#E8EFFF")], foreground=[("selected", INK)])
        style.configure("Primary.TButton", background=GOLD, foreground="#FFFFFF", padding=(18, 11), borderwidth=0, bordercolor=GOLD, font=("Microsoft YaHei UI", 9, "bold"))
        style.map("Primary.TButton", background=[("active", GOLD_DARK), ("pressed", GOLD_DARK)])
        style.configure("Danger.TButton", background=RED, foreground="#FFFFFF", padding=(16, 10), borderwidth=1, bordercolor=RED, font=("Microsoft YaHei UI", 9, "bold"))
        style.map("Danger.TButton", background=[("active", "#B91C1C")])
        style.configure("Soft.TButton", background="#FFFFFF", foreground=INK, padding=(14, 10), borderwidth=1, bordercolor=LINE, font=("Microsoft YaHei UI", 9, "bold"))
        style.map("Soft.TButton", background=[("active", "#EEF2F7")])
        style.configure("App.TEntry", fieldbackground="#FFFFFF", bordercolor=LINE, lightcolor=LINE, darkcolor=LINE, padding=9)
        style.configure("App.TSpinbox", fieldbackground="#FFFFFF", bordercolor=LINE, padding=8)
        style.configure(
            "TCombobox",
            fieldbackground="#FFFFFF",
            background="#FFFFFF",
            bordercolor=LINE,
            lightcolor=LINE,
            darkcolor=LINE,
            arrowcolor=GOLD,
            padding=7,
        )
        style.map(
            "TCombobox",
            fieldbackground=[("readonly", "#FFFFFF")],
            selectbackground=[("readonly", "#FFFFFF")],
            selectforeground=[("readonly", INK)],
        )

    def _build_background(self) -> None:
        self.background_label = tk.Label(self, bd=0, highlightthickness=0, bg=PANEL)
        self.background_label.place(x=0, y=0, relwidth=1, relheight=1)
        self.bind("<Configure>", self._schedule_background_render)
        self.after_idle(self._render_background)

    def _schedule_background_render(self, _event=None) -> None:
        if self._resize_job:
            self.after_cancel(self._resize_job)
        self._resize_job = self.after(100, self._render_background)

    def _render_background(self) -> None:
        self._resize_job = None
        image_path = ""
        if hasattr(self, "settings"):
            image_path = str(getattr(self.settings, "background_image", "") or "")
        if image_path and os.path.isfile(image_path):
            try:
                width = max(1, self.winfo_width())
                height = max(1, self.winfo_height())
                fitted = ImageOps.fit(
                    Image.open(image_path).convert("RGB"),
                    (width, height),
                    method=Image.Resampling.BILINEAR,
                )
                self._background_photo = ImageTk.PhotoImage(fitted)
                self.background_label.configure(image=self._background_photo, bg=PANEL)
            except (OSError, tk.TclError):
                self.background_label.configure(image="", bg=PANEL)
        else:
            self.background_label.configure(image="", bg=PANEL)
        self.background_label.lower()

    def _build_header(self) -> None:
        header = tk.Frame(
            self, bg=PANEL, height=52, bd=0,
            highlightthickness=1, highlightbackground=LINE,
        )
        header.place(x=0, y=0, relwidth=1)
        header.pack_propagate(False)
        self.connection_var = tk.StringVar(value=f"{tr(remaining_text(self._trial_status))} · {tr('本地模式')}")
        tk.Label(
            header, textvariable=self.connection_var, bg=PANEL, fg=MUTED,
            font=("Microsoft YaHei UI", 9), padx=24,
        ).pack(side="left")
        tk.Label(
            header, text=tr("跨境邮件营销自动化助手"), bg=PANEL, fg=MUTED,
            font=("Microsoft YaHei UI", 9),
        ).pack(side="left", padx=(0, 24))
        version_badge = tk.Frame(header, bg=GOLD, bd=0)
        version_badge.pack(side="right", padx=16, pady=13)
        tk.Label(
            version_badge, text=APP_VERSION, bg=GOLD, fg="#FFFFFF",
            padx=12, pady=4, font=("Segoe UI", 9, "bold"),
        ).pack()

    def _build_workspace(self) -> None:
        self.shell = tk.Frame(self, bg=PANEL, bd=0)
        self.shell.pack(fill="both", expand=True)
        self.lock_banner_var = tk.StringVar(value="")
        self.lock_banner = tk.Frame(self.shell, bg=RED, bd=0)
        self.lock_banner_label = tk.Label(
            self.lock_banner,
            textvariable=self.lock_banner_var,
            bg=RED,
            fg="#FFFFFF",
            cursor="hand2",
            padx=16,
            pady=8,
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        self.lock_banner_label.pack(fill="x")
        self.lock_banner_label.bind(
            "<Button-1>", lambda _event: self._select_page(self.settings_tab)
        )
        self.shell_body = tk.Frame(self.shell, bg=PANEL)
        self.shell_body.pack(fill="both", expand=True)
        sidebar = tk.Frame(self.shell_body, bg=SIDEBAR, width=224)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        brand_box = tk.Frame(sidebar, bg=SIDEBAR)
        brand_box.pack(fill="x", padx=18, pady=(22, 18))
        brand_row = tk.Frame(brand_box, bg=SIDEBAR)
        brand_row.pack(anchor="w")
        if self._header_icon:
            tk.Label(brand_row, image=self._header_icon, bg=SIDEBAR, bd=0).pack(
                side="left", padx=(0, 10)
            )
        title_box = tk.Frame(brand_row, bg=SIDEBAR)
        title_box.pack(side="left")
        tk.Label(
            title_box, text=tr("牛马邮箱"), bg=SIDEBAR, fg="#FFFFFF",
            font=(DISPLAY_FONT, 17, "bold"),
        ).pack(anchor="w")
        tk.Label(
            title_box, text=APP_VERSION, bg=SIDEBAR, fg="#7DB3FF",
            font=("Segoe UI", 8, "bold"),
        ).pack(anchor="w")
        tk.Label(
            brand_box, text=tr("外贸发信任务调度台"), bg=SIDEBAR, fg="#9DABC0",
            font=("Microsoft YaHei UI", 9),
        ).pack(anchor="w", pady=(8, 0))
        tk.Frame(brand_box, bg="#1E2A44", height=1).pack(fill="x", pady=(12, 0))
        content = tk.Frame(self.shell_body, bg=PANEL)
        content.pack(side="left", fill="both", expand=True)
        self.notebook = ttk.Notebook(content, style="Sidebar.TNotebook")
        self.notebook.pack(fill="both", expand=True, padx=0, pady=0)
        self.queue_tab = tk.Frame(self.notebook, bg=PANEL)
        self.batch_tab = tk.Frame(self.notebook, bg=PANEL)
        self.window_tab = tk.Frame(self.notebook, bg=PANEL)
        self.history_tab = tk.Frame(self.notebook, bg=PANEL)
        self.template_tab = tk.Frame(self.notebook, bg=PANEL)
        self.settings_tab = tk.Frame(self.notebook, bg=PANEL)
        self.notebook.add(self.queue_tab, text=tr("任务与草稿"))
        self.notebook.add(self.batch_tab, text=tr("10条录入"))
        self.notebook.add(self.window_tab, text=tr("窗口顺序"))
        self.notebook.add(self.history_tab, text=tr("历史记录"))
        self.notebook.add(self.template_tab, text=tr("邮件模板"))
        self.notebook.add(self.settings_tab, text=tr("设置"))
        self._nav_buttons: dict[tk.Frame, tk.Button] = {}
        tk.Label(
            sidebar, text=tr("任务流程"), bg=SIDEBAR, fg="#66758B",
            font=("Microsoft YaHei UI", 8, "bold"), anchor="w",
        ).pack(fill="x", padx=20, pady=(2, 6))
        nav_specs = [
            (self.queue_tab, tr("任务调度台")),
            (self.batch_tab, tr("导入联系人")),
            (self.template_tab, tr("邮件模板")),
            (self.window_tab, tr("发信窗口")),
            (self.history_tab, tr("发送记录")),
            (self.settings_tab, tr("系统设置")),
        ]
        for page, title in nav_specs:
            button = tk.Button(
                sidebar, text=title, command=lambda target=page: self._select_page(target),
                bg=SIDEBAR, fg="#FFFFFF", activebackground=SIDEBAR_HOVER,
                activeforeground="#FFFFFF", relief="flat", bd=0, anchor="w",
                padx=18, pady=11, font=("Microsoft YaHei UI", 10, "bold"),
                cursor="hand2",
            )
            button.pack(fill="x", padx=10, pady=2)
            self._nav_buttons[page] = button
        sidebar_footer = tk.Frame(sidebar, bg=SIDEBAR)
        sidebar_footer.pack(side="bottom", fill="x", padx=16, pady=14)
        tk.Frame(sidebar_footer, bg="#1E2A44", height=1).pack(fill="x", pady=(0, 10))
        self.connection_var = tk.StringVar(
            value=f"{tr(remaining_text(self._trial_status))} · {tr('本地模式')}"
        )
        tk.Label(
            sidebar_footer, textvariable=self.connection_var, bg=SIDEBAR, fg="#7DB3FF",
            font=("Microsoft YaHei UI", 8, "bold"),
        ).pack(anchor="w", pady=(0, 8))
        tk.Label(sidebar_footer, text=tr("本地处理 · 人工确认发送"), bg=SIDEBAR, fg="#94A3B8", font=("Microsoft YaHei UI", 8)).pack(anchor="w")
        tk.Label(sidebar_footer, text=APP_VERSION, bg=SIDEBAR, fg="#94A3B8", font=("Segoe UI", 8, "bold")).pack(anchor="w", pady=(4, 0))
        self._build_queue()
        self._build_batch_import()
        self._build_window_sequence()
        self._build_history()
        self._build_template_editor()
        self._build_settings()
        self.notebook.bind("<<NotebookTabChanged>>", lambda _event: self._sync_nav())
        self._sync_nav()
        self.status_var = tk.StringVar(value="准备就绪")
        status = tk.Label(self.shell, textvariable=self.status_var, bg="#E9EEF5", fg=MUTED, anchor="w", padx=18, pady=9, font=("Microsoft YaHei UI", 9))
        status.pack(fill="x", side="bottom")

    def _select_page(self, page) -> None:
        if (
            not self._trial_status.active
            and page in (self.batch_tab, self.window_tab)
        ):
            self.notebook.select(self.settings_tab)
            self._sync_nav()
            self._show_authorization_required()
            return
        self.notebook.select(page)
        self._sync_nav()

    def _sync_nav(self) -> None:
        if not hasattr(self, "_nav_buttons"):
            return
        selected = self.notebook.nametowidget(self.notebook.select())
        for page, button in self._nav_buttons.items():
            active = page is selected
            button.configure(
                bg=GOLD if active else SIDEBAR,
                activebackground=GOLD_DARK if active else SIDEBAR_HOVER,
            )
        if hasattr(self, "history_tab") and selected is self.history_tab:
            self.refresh_history(self._all_rows or None)
        if hasattr(self, "window_tab") and selected is self.window_tab:
            self._refresh_window_template_options()

    def _card(self, parent, **grid_options) -> tk.Frame:
        card = tk.Frame(parent, bg=PAPER, bd=0, highlightthickness=1, highlightbackground=LINE)
        if grid_options:
            card.grid(**grid_options)
        return card

    def _rebuild_ui(self) -> None:
        """Recreate the whole window after theme/skin/language changes."""
        self.config(cursor="watch")
        self.update_idletasks()
        try:
            self.unbind_all("<MouseWheel>")
            for child in self.winfo_children():
                child.destroy()
            apply_theme(self.settings)
            self._configure_styles()
            self._build_background()
            self._operation_buttons = []
            self._build_workspace()
            self._set_operation_busy(False)
            self.after_idle(self.refresh)
            self._translate_widgets()
            self._apply_authorization_state()
        except Exception as exc:
            logger.exception("重建界面失败")
            messagebox.showerror("应用主题失败", f"界面重建失败：{exc}\n\n请重启软件。")
        finally:
            self.config(cursor="")

    def _enable_canvas_wheel(self, canvas, frame, axis: str = "y") -> None:
        """Scroll a canvas region with the wheel even over its child widgets.

        Uses sign-based steps so precision mice/trackpads (small delta values)
        still scroll instead of being rounded down to zero.
        """
        active = {"on": False}

        def _enter(_event) -> None:
            active["on"] = True

        def _leave(_event) -> None:
            active["on"] = False

        def _wheel(event):
            if not active["on"]:
                return None
            if isinstance(event.widget, (tk.Text, ttk.Treeview)):
                return None
            steps = max(1, abs(event.delta) // 120)
            direction = -steps if event.delta > 0 else steps
            if axis == "x":
                canvas.xview_scroll(direction, "units")
            else:
                canvas.yview_scroll(direction, "units")
            return "break"

        canvas.bind("<Enter>", _enter)
        canvas.bind("<Leave>", _leave)
        frame.bind("<Enter>", _enter)
        frame.bind("<Leave>", _leave)
        self.bind_all("<MouseWheel>", _wheel, add="+")

    def _build_queue(self) -> None:
        self.queue_tab.grid_columnconfigure(0, weight=1)
        self.queue_tab.grid_rowconfigure(2, weight=1)
        self.task_filter = "active"

        switcher = self._card(self.queue_tab, row=0, column=0, sticky="ew", padx=20, pady=(18, 6))
        mission = tk.Frame(switcher, bg=PAPER)
        mission.pack(fill="x", padx=18, pady=(14, 8))
        mission_copy = tk.Frame(mission, bg=PAPER)
        mission_copy.pack(side="left")
        tk.Label(
            mission_copy, text=tr("本轮发信任务"), bg=PAPER, fg=INK,
            font=(DISPLAY_FONT, 18, "bold"),
        ).pack(anchor="w")
        tk.Label(
            mission_copy, text=tr("按流程准备联系人、生成内容并填写 Gmail 草稿"),
            bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 9),
        ).pack(anchor="w", pady=(3, 0))
        provider = tk.Frame(mission, bg=PAPER)
        provider.pack(side="right")
        tk.Label(
            provider, text=tr("发信窗口"), bg=PAPER, fg=MUTED,
            font=("Microsoft YaHei UI", 8, "bold"),
        ).pack(side="left", padx=(0, 8))
        provider_name = BROWSER_PROVIDER_NAMES.get(
            self.settings.browser_provider,
            BROWSER_PROVIDER_NAMES["morelogin"],
        )
        self.browser_provider_var = tk.StringVar(value=provider_name)
        for name in BROWSER_PROVIDER_NAMES.values():
            tk.Radiobutton(
                provider,
                text=name,
                value=name,
                variable=self.browser_provider_var,
                command=self.switch_browser_provider,
                indicatoron=False,
                bg="#EDF1F6",
                fg=INK,
                selectcolor="#E8EFFF",
                activebackground="#E8EFFF",
                activeforeground=INK,
                relief="flat",
                bd=0,
                padx=14,
                pady=6,
                cursor="hand2",
                font=("Microsoft YaHei UI", 8, "bold"),
            ).pack(side="left", padx=2)
        self.test_connection_button = ttk.Button(
            provider,
            text=tr("测试当前连接"),
            style="Soft.TButton",
            command=self.test_browser_connection,
        )
        self.test_connection_button.pack(side="left", padx=(8, 0))

        route = tk.Frame(switcher, bg="#F7F9FC")
        route.pack(fill="x", padx=1, pady=(0, 1))
        route_inner = tk.Frame(route, bg="#F7F9FC")
        route_inner.pack(fill="x", padx=18, pady=10)
        route_steps = [
            ("1", tr("导入联系人"), tr("已就绪"), self.batch_tab),
            ("2", tr("配置模板"), tr("已就绪"), self.template_tab),
            ("3", tr("生成并填写"), tr("当前步骤"), self.queue_tab),
            ("4", tr("确认结果"), tr("待开始"), self.history_tab),
        ]
        for index, (number, label, state, target) in enumerate(route_steps):
            step = tk.Frame(route_inner, bg="#F7F9FC", cursor="hand2")
            step.pack(side="left", fill="x", expand=True)
            dot_color = GOLD if index == 2 else (GREEN if index < 2 else "#CBD5E1")
            dot = tk.Label(
                step, text=number, bg=dot_color, fg="#FFFFFF",
                width=2, pady=3, font=("Segoe UI", 8, "bold"), cursor="hand2",
            )
            dot.pack(side="left")
            copy = tk.Frame(step, bg="#F7F9FC", cursor="hand2")
            copy.pack(side="left", padx=(8, 0))
            name = tk.Label(copy, text=label, bg="#F7F9FC", fg=INK if index <= 2 else MUTED,
                            font=("Microsoft YaHei UI", 9, "bold"), cursor="hand2")
            name.pack(anchor="w")
            state_label = tk.Label(copy, text=state, bg="#F7F9FC",
                                   fg=GREEN if index < 2 else (GOLD if index == 2 else MUTED),
                                   font=("Microsoft YaHei UI", 7), cursor="hand2")
            state_label.pack(anchor="w")
            for widget in (step, dot, copy, name, state_label):
                widget.bind("<Button-1>", lambda _event, page=target: self._select_page(page))
            if index < len(route_steps) - 1:
                tk.Frame(route_inner, bg=dot_color, height=2, width=32).pack(
                    side="left", padx=10
                )

        stats = tk.Frame(self.queue_tab, bg=PANEL)
        stats.grid(row=1, column=0, sticky="ew", padx=15, pady=(6, 7))
        for i in range(4):
            stats.grid_columnconfigure(i, weight=1)
        self.stat_vars = {key: tk.StringVar(value="0") for key in ("all", "new", "ready", "drafted")}
        stat_specs = [
            ("all", tr("本轮任务"), GOLD),
            ("new", tr("等待生成"), "#94A3B8"),
            ("ready", tr("等待确认"), AMBER),
            ("drafted", tr("草稿已填写"), GREEN),
        ]
        for idx, (key, label, color) in enumerate(stat_specs):
            card = self._card(stats, row=0, column=idx, sticky="ew", padx=5)
            accent = tk.Frame(card, bg=color, width=5)
            accent.pack(side="left", fill="y")
            body = tk.Frame(card, bg=PAPER)
            body.pack(side="left", padx=14, pady=9)
            count_label = tk.Label(body, textvariable=self.stat_vars[key], bg=PAPER, fg=INK, font=(DISPLAY_FONT, 19, "bold"))
            count_label.pack(anchor="w")
            name_label = tk.Label(body, text=label, bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8))
            name_label.pack(anchor="w")
            for widget in (card, accent, body, count_label, name_label):
                widget.configure(cursor="hand2")
                widget.bind("<Button-1>", lambda _event, value=key: self._set_task_filter(value))
        # Counts remain live for the persistent dispatch summary; the legacy
        # metric strip is intentionally hidden to keep one operational spine.
        stats.grid_remove()

        actions_shell = tk.Frame(self.queue_tab, bg=PANEL)
        actions_shell.grid(row=2, column=0, sticky="ew", padx=19, pady=5)
        actions_canvas = tk.Canvas(
            actions_shell, bg=PANEL, highlightthickness=0, height=46
        )
        actions_hbar = ttk.Scrollbar(
            actions_shell, orient="horizontal", command=actions_canvas.xview
        )
        actions = tk.Frame(actions_canvas, bg=PANEL)
        actions_window = actions_canvas.create_window(
            (0, 0), window=actions, anchor="nw"
        )
        actions_canvas.configure(xscrollcommand=actions_hbar.set)
        actions_canvas.pack(side="top", fill="x")
        actions_hbar.pack(side="bottom", fill="x")
        actions.bind(
            "<Configure>",
            lambda _event: actions_canvas.configure(
                scrollregion=actions_canvas.bbox("all")
            ),
        )
        actions_canvas.bind(
            "<Configure>",
            lambda event: actions_canvas.itemconfigure(
                actions_window,
                width=max(event.width, actions.winfo_reqwidth()),
            ),
        )

        def _on_actions_wheel(event) -> str:
            actions_canvas.xview_scroll(-1 if event.delta > 0 else 1, "units")
            return "break"

        def _bind_actions_wheel(widget) -> None:
            widget.bind("<MouseWheel>", _on_actions_wheel)
            for child in widget.winfo_children():
                _bind_actions_wheel(child)

        self.ungenerated_hint_var = tk.StringVar(value="")
        tk.Label(
            actions,
            textvariable=self.ungenerated_hint_var,
            bg=PANEL,
            fg=RED,
            font=("Microsoft YaHei UI", 8, "bold"),
        ).pack(side="left", padx=(0, 10))
        self.generate_ungenerated_button = ttk.Button(
            actions,
            text=tr("生成未生成任务"),
            style="Soft.TButton",
            command=self.generate_ungenerated_tasks,
        )
        self.generate_ungenerated_button.pack(side="left", padx=(0, 6))
        ttk.Button(actions, text=tr("＋ 导入联系人"), style="Soft.TButton", command=lambda: self._select_page(self.batch_tab)).pack(side="left", padx=(0, 6))
        self.generate_selected_button = ttk.Button(actions, text=tr("1  本地生成所选"), style="Primary.TButton", command=self.generate_selected_tasks)
        self.generate_selected_button.pack(side="left", padx=(0, 6))
        self.open_drafts_button = ttk.Button(actions, text=tr("2  填写 Gmail 草稿"), style="Primary.TButton", command=lambda: self.open_selected_drafts(False))
        self.open_drafts_button.pack(side="left", padx=(0, 6))
        self.wait_send_button = ttk.Button(actions, text=tr("填写并自动发送"), style="Primary.TButton", command=lambda: self.open_selected_drafts(True))
        self.wait_send_button.pack(side="left", padx=(0, 6))
        self.cancel_operation_button = ttk.Button(actions, text=tr("停止当前任务"), style="Soft.TButton", command=self.cancel_current_operation)
        self.cancel_operation_button.pack(side="left")
        self.retry_failed_button = ttk.Button(actions, text=tr("重试失败草稿"), style="Soft.TButton", command=self.retry_failed_drafts)
        self.retry_failed_button.pack(side="left", padx=(6, 0))
        self.retry_failed_button.state(["disabled"])
        self.mark_sent_button = ttk.Button(actions, text=tr("标记已发送"), style="Soft.TButton", command=self.mark_selected_sent)
        self.mark_sent_button.pack(side="left", padx=(6, 0))
        self.unmark_sent_button = ttk.Button(actions, text=tr("撤销标记"), style="Soft.TButton", command=self.unmark_selected_sent)
        self.unmark_sent_button.pack(side="left", padx=(6, 0))
        ttk.Button(actions, text=tr("删除"), style="Danger.TButton", command=self.delete_selected_queue_tasks).pack(side="right", padx=(6, 0))
        self._operation_buttons.extend(
            [
                self.generate_selected_button,
                self.open_drafts_button,
                self.wait_send_button,
                self.generate_ungenerated_button,
            ]
        )
        _bind_actions_wheel(actions)
        # Recovery commands stay instantiated for operation-state management,
        # but are surfaced contextually through the compact menu below.
        actions_shell.grid_remove()

        filters = tk.Frame(self.queue_tab, bg=PANEL)
        filters.grid(row=1, column=0, sticky="ew", padx=20, pady=(10, 7))
        self.selection_count_var = tk.StringVar(value=trf("已选择 {count} 条", count=0))
        tk.Label(filters, textvariable=self.selection_count_var, bg=PANEL, fg=INK, font=("Microsoft YaHei UI", 8, "bold")).pack(side="left")
        self.select_all_button = ttk.Button(filters, text=tr("全选当前结果"), style="Soft.TButton", command=self.select_all_queue_tasks)
        self.select_all_button.pack(side="left", padx=8)
        self._operation_buttons.append(self.select_all_button)
        ttk.Button(filters, text=tr("刷新"), style="Soft.TButton", command=self.refresh).pack(side="right")
        more_actions = tk.Menubutton(
            filters, text=tr("更多操作"), bg=PAPER, fg=INK,
            activebackground="#EEF2F7", activeforeground=INK,
            relief="solid", bd=1, padx=12, pady=8,
            font=("Microsoft YaHei UI", 9, "bold"), cursor="hand2",
        )
        more_menu = tk.Menu(more_actions, tearoff=False)
        more_menu.add_command(label=tr("生成全部待处理任务"), command=self.generate_ungenerated_tasks)
        more_menu.add_command(label=tr("填写并跟踪发送"), command=lambda: self.open_selected_drafts(True))
        more_menu.add_separator()
        more_menu.add_command(label=tr("重试失败草稿"), command=self.retry_failed_drafts)
        more_menu.add_command(label=tr("停止当前任务"), command=self.cancel_current_operation)
        more_menu.add_separator()
        more_menu.add_command(label=tr("标记已发送"), command=self.mark_selected_sent)
        more_menu.add_command(label=tr("撤销发送标记"), command=self.unmark_selected_sent)
        more_menu.add_command(label=tr("删除所选任务"), command=self.delete_selected_queue_tasks)
        more_actions.configure(menu=more_menu)
        more_actions.pack(side="right", padx=(8, 0))
        ttk.Button(filters, text=tr("今日统计"), style="Soft.TButton", command=self.show_daily_stats).pack(side="right", padx=(8, 0))
        ttk.Button(filters, text=tr("窗口状态"), style="Soft.TButton", command=self.show_window_status).pack(side="right", padx=(8, 0))
        self.task_search_var = tk.StringVar()
        self.task_search_var.trace_add("write", lambda *_args: self._schedule_filter_refresh())
        search = ttk.Entry(filters, textvariable=self.task_search_var, style="App.TEntry", width=28)
        search.pack(side="right", padx=(8, 0))
        tk.Label(filters, text=tr("搜索姓名 / 城市 / 邮箱"), bg=PANEL, fg=MUTED, font=("Microsoft YaHei UI", 8)).pack(side="right")

        split = tk.PanedWindow(self.queue_tab, orient="horizontal", bg=PANEL, bd=0, sashwidth=8, sashrelief="flat")
        split.grid(row=2, column=0, sticky="nsew", padx=20, pady=(2, 14))
        table_card = self._card(split)
        preview_card = self._card(split)
        split.add(table_card, minsize=650, stretch="always")
        split.add(preview_card, minsize=310, stretch="always")

        table_heading = tk.Frame(table_card, bg=PAPER)
        table_heading.pack(fill="x", padx=14, pady=(12, 10))
        heading_copy = tk.Frame(table_heading, bg=PAPER)
        heading_copy.pack(side="left")
        tk.Label(heading_copy, text=tr("联系人任务清单"), bg=PAPER, fg=INK,
                 font=("Microsoft YaHei UI", 11, "bold")).pack(anchor="w")
        tk.Label(heading_copy, text=tr("还没有任务时，请先导入 XLSX / CSV，或手动录入联系人"),
                 bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8)).pack(anchor="w", pady=(2, 0))
        ttk.Button(table_heading, text=tr("导入或录入联系人"), style="Primary.TButton",
                   command=lambda: self._select_page(self.batch_tab)).pack(side="right")

        columns = ("id", "profile", "name", "location", "email", "status")
        self.tree = ttk.Treeview(table_card, columns=columns, show="headings", style="App.Treeview", selectmode="extended")
        labels = tuple(tr(label) for label in ("ID", "窗口选择", "姓名（点击编辑）", "地点（点击编辑）", "邮箱", "状态"))
        widths = (42, 112, 105, 125, 185, 170)
        for col, label, width in zip(columns, labels, widths):
            self.tree.heading(col, text=label)
            self.tree.column(col, width=width, minwidth=45, anchor="w")
        scroll = ttk.Scrollbar(table_card, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.tree.bind("<<TreeviewSelect>>", lambda _e: self._on_task_selection())
        self.tree.bind("<ButtonRelease-1>", self._handle_task_table_click, add="+")
        self.tree.bind("<Delete>", lambda _event: self.delete_selected_queue_tasks())
        self.tree.bind("<Control-a>", self._select_all_shortcut)

        dispatch = tk.Frame(preview_card, bg="#0F1C31")
        dispatch.pack(fill="x", padx=1, pady=(1, 10))
        tk.Label(dispatch, text=tr("发送签派摘要"), bg="#0F1C31", fg="#FFFFFF",
                 font=(DISPLAY_FONT, 14, "bold")).pack(anchor="w", padx=14, pady=(14, 2))
        tk.Label(dispatch, textvariable=self.selection_count_var, bg="#0F1C31", fg="#9FB6D8",
                 font=("Microsoft YaHei UI", 8)).pack(anchor="w", padx=14, pady=(0, 10))
        checklist = tk.Frame(dispatch, bg="#0F1C31")
        checklist.pack(fill="x", padx=14)
        dispatch_rows = [
            (tr("联系人已导入"), self.stat_vars["all"], GREEN),
            (tr("等待生成内容"), self.stat_vars["new"], "#94A3B8"),
            (tr("等待人工确认"), self.stat_vars["ready"], AMBER),
            (tr("Gmail 草稿就绪"), self.stat_vars["drafted"], SKY),
        ]
        for label, value_var, color in dispatch_rows:
            row = tk.Frame(checklist, bg="#0F1C31")
            row.pack(fill="x", pady=3)
            tk.Label(row, text="●", bg="#0F1C31", fg=color,
                     font=("Segoe UI", 8)).pack(side="left")
            tk.Label(row, text=label, bg="#0F1C31", fg="#DCE6F4",
                     font=("Microsoft YaHei UI", 8)).pack(side="left", padx=(7, 0))
            tk.Label(row, textvariable=value_var, bg="#0F1C31", fg="#FFFFFF",
                     font=("Segoe UI", 9, "bold")).pack(side="right")
        dispatch_actions = tk.Frame(dispatch, bg="#0F1C31")
        dispatch_actions.pack(fill="x", padx=14, pady=(12, 14))
        ttk.Button(dispatch_actions, text=tr("生成所选内容"), style="Primary.TButton",
                   command=self.generate_selected_tasks).pack(fill="x")
        ttk.Button(dispatch_actions, text=tr("填写 Gmail 草稿"), style="Soft.TButton",
                   command=lambda: self.open_selected_drafts(False)).pack(fill="x", pady=(6, 0))

        tk.Label(preview_card, text=tr("邮件预览"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 11, "bold")).pack(anchor="w", padx=14, pady=(2, 2))
        tk.Label(preview_card, text=tr("发送前请在 Gmail 中逐封核对"), bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8)).pack(anchor="w", padx=14, pady=(0, 8))
        self.preview = tk.Text(preview_card, wrap="word", state="disabled", bg="#FFFFFF", fg=INK, relief="flat", highlightthickness=1, highlightbackground=LINE, padx=12, pady=12, font=("Microsoft YaHei UI", 9), spacing1=2, spacing3=4)
        self.preview.pack(fill="both", expand=True, padx=12, pady=(0, 6))
        picker = tk.Frame(preview_card, bg="#F8FAFD", highlightthickness=1, highlightbackground=LINE)
        picker.pack(fill="x", padx=12, pady=(0, 6))
        tk.Label(picker, text=tr("选择发送窗口"), bg="#F8FAFD", fg=INK, font=("Microsoft YaHei UI", 8, "bold")).pack(side="left", padx=(10, 6), pady=8)
        self.profile_assign_var = tk.IntVar(value=1)
        ttk.Spinbox(picker, from_=1, to=999, textvariable=self.profile_assign_var, width=7, style="App.TSpinbox").pack(side="left", pady=5)
        self.window_picker_var = tk.StringVar()
        self.window_picker_combo = ttk.Combobox(
            picker,
            textvariable=self.window_picker_var,
            state="readonly",
            width=16,
            postcommand=self.auto_populate_picker_quiet,
        )
        self.window_picker_combo.pack(side="left", padx=(4, 0), pady=5)
        self.window_picker_combo.bind(
            "<<ComboboxSelected>>", self._on_window_picker_selected
        )
        ttk.Button(
            picker,
            text=tr("自动识别"),
            style="Soft.TButton",
            command=self.refresh_window_picker,
        ).pack(side="left", padx=(4, 0), pady=5)
        self.profile_assign_button = ttk.Button(picker, text=tr("确认浏览器窗口"), style="Soft.TButton", command=self.assign_profile_to_selected)
        self.profile_assign_button.pack(side="right", padx=8, pady=5)
        self.profile_assign_note_var = tk.StringVar(value=tr("请选择任务"))
        tk.Label(picker, textvariable=self.profile_assign_note_var, bg="#F8FAFD", fg=MUTED, font=("Microsoft YaHei UI", 8)).pack(side="right", padx=4)

        editor = tk.Frame(preview_card, bg=PAPER, highlightthickness=1, highlightbackground=LINE)
        editor.pack(fill="x", padx=12, pady=(0, 6))
        tk.Label(editor, text=tr("手动填写或编辑资料"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 8, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=9, pady=(7, 4))
        tk.Label(editor, text=tr("姓名/称呼"), bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8)).grid(row=1, column=0, sticky="w", padx=(9, 5), pady=3)
        self.manual_name_var = tk.StringVar()
        ttk.Entry(editor, textvariable=self.manual_name_var, style="App.TEntry").grid(row=1, column=1, sticky="ew", padx=(0, 9), pady=3)
        tk.Label(editor, text=tr("城市/城市地区"), bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8)).grid(row=2, column=0, sticky="w", padx=(9, 5), pady=3)
        self.manual_location_var = tk.StringVar()
        ttk.Entry(editor, textvariable=self.manual_location_var, style="App.TEntry").grid(row=2, column=1, sticky="ew", padx=(0, 9), pady=3)
        tk.Label(editor, text=tr("发件人名字"), bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8)).grid(row=3, column=0, sticky="w", padx=(9, 5), pady=3)
        self.manual_sender_var = tk.StringVar()
        ttk.Entry(editor, textvariable=self.manual_sender_var, style="App.TEntry").grid(row=3, column=1, sticky="ew", padx=(0, 9), pady=3)
        tk.Label(editor, text=tr("自定义变量"), bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8)).grid(row=4, column=0, sticky="nw", padx=(9, 5), pady=3)
        self.manual_custom_frame = tk.Frame(editor, bg=PAPER)
        self.manual_custom_frame.grid(row=4, column=1, sticky="ew", padx=(0, 9), pady=3)
        self._render_manual_custom_fields()
        ttk.Button(editor, text=tr("保存资料并重新生成"), style="Primary.TButton", command=self.save_manual_profile).grid(row=5, column=1, sticky="e", padx=9, pady=(5, 8))
        editor.grid_columnconfigure(1, weight=1)

    def _render_manual_custom_fields(self) -> None:
        for child in self.manual_custom_frame.winfo_children():
            child.destroy()
        self.manual_custom_vars: dict[str, tk.StringVar] = {}
        keys = list(self.settings.custom_variable_keys)
        if not keys:
            tk.Label(
                self.manual_custom_frame,
                text=tr("暂无自定义变量，可在模板页添加"),
                bg=PAPER,
                fg=MUTED,
                font=("Microsoft YaHei UI", 8),
            ).pack(anchor="w")
            return
        for key in keys:
            line = tk.Frame(self.manual_custom_frame, bg=PAPER)
            line.pack(fill="x", pady=1)
            tk.Label(
                line,
                text=f"{{{key}}}",
                bg=PAPER,
                fg=GOLD_DARK,
                font=("Segoe UI", 8, "bold"),
                width=12,
                anchor="w",
            ).pack(side="left")
            value_var = tk.StringVar()
            self.manual_custom_vars[key] = value_var
            ttk.Entry(line, textvariable=value_var, style="App.TEntry").pack(
                side="left", fill="x", expand=True
            )
        self._translate_widgets()

    def _visible_entry_fields(self) -> list[tuple[str, str, int]]:
        fields: list[tuple[str, str, int]] = []
        if "first_name" not in self.settings.hidden_system_variables:
            fields.append(("name", "名字", 2))
        if "location" not in self.settings.hidden_system_variables:
            fields.append(("location", "地区", 2))
        fields.append(("email", "邮箱地址", 3))
        return fields

    def _build_batch_import(self) -> None:
        self.batch_tab.grid_columnconfigure(0, weight=1)
        self.batch_tab.grid_rowconfigure(1, weight=1)
        header = tk.Frame(self.batch_tab, bg=PANEL)
        header.grid(row=0, column=0, sticky="ew", padx=22, pady=(18, 8))
        tk.Label(header, text=tr("联系人录入"), bg=PANEL, fg=INK, font=("Microsoft YaHei UI", 14, "bold")).pack(side="left")
        tk.Label(
            header,
            text=trf(
                "手动填写后完全在本机生成 · 最多 {count} 行 · 自定义变量按需添加",
                count=MAX_CONTACT_ROWS,
            ),
            bg=PANEL,
            fg=GREEN,
            font=("Microsoft YaHei UI", 8, "bold"),
        ).pack(side="left", padx=14)

        card = self._card(self.batch_tab, row=1, column=0, sticky="nsew", padx=22, pady=(0, 18))
        card.grid_rowconfigure(1, weight=1)
        card.grid_columnconfigure(0, weight=1)

        self._entry_fields = self._visible_entry_fields()
        headings_frame = tk.Frame(card, bg="#E8EFFA")
        headings_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=1, pady=(1, 4))
        for column, (_key, _label, weight) in enumerate(self._entry_fields, start=1):
            headings_frame.grid_columnconfigure(column, weight=weight)
        self.batch_headings_frame = headings_frame
        self._rebuild_batch_headings()
        self._translate_widgets()

        rows_canvas = tk.Canvas(card, bg=PAPER, highlightthickness=0, bd=0)
        rows_scroll = ttk.Scrollbar(card, orient="vertical", command=rows_canvas.yview)
        rows_canvas.configure(yscrollcommand=rows_scroll.set)
        rows_canvas.grid(row=1, column=0, sticky="nsew")
        rows_scroll.grid(row=1, column=1, sticky="ns")
        rows_frame = tk.Frame(rows_canvas, bg=PAPER)
        for column, (_key, _label, weight) in enumerate(self._entry_fields, start=1):
            rows_frame.grid_columnconfigure(column, weight=weight)
        rows_window = rows_canvas.create_window((0, 0), window=rows_frame, anchor="nw")
        rows_frame.bind(
            "<Configure>",
            lambda _event: rows_canvas.configure(scrollregion=rows_canvas.bbox("all")),
        )
        rows_canvas.bind(
            "<Configure>",
            lambda event: rows_canvas.itemconfigure(rows_window, width=event.width),
        )

        self._enable_canvas_wheel(rows_canvas, rows_frame)
        self.batch_rows_frame = rows_frame
        self._build_batch_rows()
        footer = tk.Frame(card, bg=PAPER)
        footer.grid(row=2, column=0, columnspan=2, sticky="ew", padx=12, pady=(10, 14))
        ttk.Button(footer, text=tr("导入 XLSX"), style="Soft.TButton", command=self.import_contacts_xlsx).pack(side="left", padx=(0, 6))
        ttk.Button(footer, text=tr("下载 XLSX 模板"), style="Soft.TButton", command=self.download_xlsx_template).pack(side="left", padx=(0, 6))
        ttk.Button(footer, text=tr("＋ 添加一行"), style="Soft.TButton", command=self.add_contact_row).pack(side="left", padx=(0, 6))
        ttk.Button(footer, text=tr("清空输入"), style="Soft.TButton", command=self.clear_contact_rows).pack(side="left")
        ttk.Button(footer, text=tr("将填写内容全部加入队列"), style="Primary.TButton", command=self.add_contacts).pack(side="right")
        self.generate_all_contacts_button = ttk.Button(footer, text=tr("同时生成已填写邮件"), style="Soft.TButton", command=self.generate_all_contacts)
        self.generate_all_contacts_button.pack(side="right", padx=(0, 8))
        self._operation_buttons.extend(self.batch_generate_buttons)
        self._operation_buttons.append(self.generate_all_contacts_button)

    def _build_batch_rows(self) -> None:
        for child in self.batch_rows_frame.winfo_children():
            child.destroy()
        fields = self._visible_entry_fields()
        self._entry_fields = fields
        for column in range(12):
            self.batch_rows_frame.grid_columnconfigure(column, weight=0)
        for column, (_key, _label, weight) in enumerate(fields, start=1):
            self.batch_rows_frame.grid_columnconfigure(column, weight=weight)
        self.contact_rows: list[tuple[tk.StringVar, tk.StringVar, tk.StringVar, dict[str, tk.StringVar]]] = []
        self.batch_generate_buttons: list[ttk.Button] = []
        if not hasattr(self, "_contact_row_count"):
            self._contact_row_count = BATCH_CONTACT_ROWS
        variable_keys = list(self.settings.custom_variable_keys)
        rows_per_record = 2 if variable_keys else 1
        template_rows = 1 + (1 if variable_keys else 0)
        action_column = len(fields) + 1
        total_columns = action_column + 2
        self._build_template_row(fields, variable_keys)
        for index in range(self._contact_row_count):
            name_var = tk.StringVar()
            location_var = tk.StringVar()
            email_var = tk.StringVar()
            custom_vars = {
                key: tk.StringVar(value=str(self.settings.custom_variables.get(key, "")))
                for key in variable_keys
            }
            self.contact_rows.append((name_var, location_var, email_var, custom_vars))
            top_row = template_rows + index * rows_per_record
            tk.Label(self.batch_rows_frame, text=str(index + 1), bg=PAPER, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(row=top_row, column=0, padx=8, pady=5)
            field_vars = {
                "name": name_var,
                "location": location_var,
                "email": email_var,
            }
            for column, (key, _label, _weight) in enumerate(fields, start=1):
                ttk.Entry(
                    self.batch_rows_frame,
                    textvariable=field_vars[key],
                    style="App.TEntry",
                ).grid(row=top_row, column=column, sticky="ew", padx=6, pady=4)
            row_generate_button = ttk.Button(
                self.batch_rows_frame,
                text=tr("生成邮件"),
                style="Soft.TButton",
                command=lambda row_index=index: self.generate_contact_row(row_index),
            )
            row_generate_button.grid(row=top_row, column=action_column, padx=(2, 10), pady=4)
            self.batch_generate_buttons.append(row_generate_button)
            ttk.Button(
                self.batch_rows_frame,
                text=tr("删除"),
                style="Danger.TButton",
                command=lambda row_index=index: self.delete_contact_row(row_index),
            ).grid(row=top_row, column=action_column + 1, padx=(0, 10), pady=4)

            if not variable_keys:
                continue
            variable_box = tk.Frame(self.batch_rows_frame, bg="#F8FAFD", highlightthickness=1, highlightbackground=LINE)
            variable_box.grid(row=top_row + 1, column=0, columnspan=total_columns, sticky="ew", padx=6, pady=(0, 6))
            tk.Label(
                variable_box,
                text=tr("自定义变量"),
                bg="#F8FAFD",
                fg=GREEN,
                font=("Microsoft YaHei UI", 8, "bold"),
            ).grid(row=0, column=0, sticky="w", padx=(8, 6), pady=5)
            for column, key in enumerate(variable_keys, start=1):
                variable_box.grid_columnconfigure(column, weight=1)
                cell = tk.Frame(variable_box, bg="#F8FAFD")
                cell.grid(row=0, column=column, sticky="ew", padx=3)
                tk.Label(
                    cell,
                    text=f"{{{key}}}",
                    bg="#F8FAFD",
                    fg=GOLD_DARK,
                    font=("Segoe UI", 8, "bold"),
                ).pack(anchor="w")
                ttk.Entry(cell, textvariable=custom_vars[key], style="App.TEntry").pack(fill="x")
        self._rebuild_batch_headings()

    def _rebuild_batch_headings(self) -> None:
        if not hasattr(self, "batch_headings_frame"):
            return
        for child in self.batch_headings_frame.winfo_children():
            child.destroy()
        headings = (
            [tr("序号")]
            + [tr(label) for _key, label, _weight in self._entry_fields]
            + [tr("本地生成"), tr("删除")]
        )
        for column, text in enumerate(headings):
            tk.Label(
                self.batch_headings_frame,
                text=text,
                bg="#E8EFFA",
                fg=INK,
                font=("Microsoft YaHei UI", 9, "bold"),
                padx=8,
                pady=9,
            ).grid(row=0, column=column, sticky="ew", padx=1)
        self._translate_widgets()

    def _build_template_row(
        self,
        fields: list[tuple[str, str, int]],
        variable_keys: list[str],
    ) -> None:
        bg = "#F1F5F9"
        action_column = len(fields) + 1
        total_columns = action_column + 2
        tk.Label(
            self.batch_rows_frame,
            text=tr("示例"),
            bg=bg,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            padx=8,
            pady=5,
        ).grid(row=0, column=0, sticky="w")
        english_ui = current_language() == "en"
        sample_values = {
            "name": "John" if english_ui else "张三",
            "location": "Seattle",
            "email": "john@example.com" if english_ui else "zhangsan@example.com",
        }
        for column, (key, _label, _weight) in enumerate(fields, start=1):
            entry = ttk.Entry(self.batch_rows_frame, style="App.TEntry")
            entry.insert(0, sample_values[key])
            entry.configure(state="disabled")
            entry.grid(row=0, column=column, sticky="ew", padx=6, pady=4)
        tk.Label(
            self.batch_rows_frame,
            text=tr("模板行（仅展示，不会导入）"),
            bg=bg,
            fg=GOLD_DARK,
            font=("Microsoft YaHei UI", 8, "bold"),
        ).grid(row=0, column=action_column, columnspan=2, sticky="w", padx=8)

        if not variable_keys:
            return
        variable_box = tk.Frame(
            self.batch_rows_frame,
            bg=bg,
            highlightthickness=1,
            highlightbackground=LINE,
        )
        variable_box.grid(row=1, column=0, columnspan=total_columns, sticky="ew", padx=6, pady=(0, 6))
        tk.Label(
            variable_box,
            text=tr("自定义变量"),
            bg=bg,
            fg=GREEN,
            font=("Microsoft YaHei UI", 8, "bold"),
        ).grid(row=0, column=0, sticky="w", padx=(8, 6), pady=5)
        sample_vars = {"custom_1": "酒店推荐", "custom_2": "美食", "custom_3": "安全"}
        for column, key in enumerate(variable_keys, start=1):
            variable_box.grid_columnconfigure(column, weight=1)
            cell = tk.Frame(variable_box, bg=bg)
            cell.grid(row=0, column=column, sticky="ew", padx=3)
            tk.Label(
                cell,
                text=f"{{{key}}}",
                bg=bg,
                fg=GOLD_DARK,
                font=("Segoe UI", 8, "bold"),
            ).pack(anchor="w")
            entry = ttk.Entry(cell, style="App.TEntry")
            entry.insert(0, sample_vars.get(key, ""))
            entry.configure(state="disabled")
            entry.pack(fill="x")

    def _capture_contact_values(self) -> list[tuple[str, str, str, dict[str, str]]]:
        captured: list[tuple[str, str, str, dict[str, str]]] = []
        for name_var, location_var, email_var, custom_vars in self.contact_rows:
            captured.append(
                (
                    name_var.get(),
                    location_var.get(),
                    email_var.get(),
                    {key: var.get() for key, var in custom_vars.items()},
                )
            )
        return captured

    def _restore_contact_values(self, captured) -> None:
        for index, (name, location, email, custom_values) in enumerate(captured):
            if index >= len(self.contact_rows):
                break
            name_var, location_var, email_var, custom_vars = self.contact_rows[index]
            name_var.set(name)
            location_var.set(location)
            email_var.set(email)
            for key, value in custom_values.items():
                if key in custom_vars:
                    custom_vars[key].set(value)

    def _refresh_batch_rows(self) -> None:
        self._operation_buttons = [
            button
            for button in getattr(self, "_operation_buttons", [])
            if button.winfo_exists()
        ]
        self._build_batch_rows()
        self._operation_buttons.extend(self.batch_generate_buttons)

    def _rebuild_contact_rows(self) -> None:
        if hasattr(self, "batch_rows_frame"):
            self._refresh_batch_rows()

    def add_contact_row(self) -> None:
        if self._contact_row_count >= MAX_CONTACT_ROWS:
            messagebox.showwarning(
                "已达到上限", f"最多只能添加 {MAX_CONTACT_ROWS} 行。"
            )
            return
        captured = self._capture_contact_values()
        self._contact_row_count += 1
        self._refresh_batch_rows()
        self._restore_contact_values(captured)
        self._set_status(
            f"已添加一行，当前 {self._contact_row_count} 行（最多 {MAX_CONTACT_ROWS} 行）"
        )

    def delete_contact_row(self, row_index: int) -> None:
        if self._contact_row_count <= 1:
            self._set_status("至少保留一行")
            return
        if not (0 <= row_index < self._contact_row_count):
            return
        captured = self._capture_contact_values()
        captured.pop(row_index)
        self._contact_row_count -= 1
        self._refresh_batch_rows()
        self._restore_contact_values(captured)
        self._set_status(
            f"已删除第 {row_index + 1} 行，当前 {self._contact_row_count} 行"
        )

    def _ensure_contact_rows(self, needed: int) -> None:
        if needed <= self._contact_row_count:
            return
        captured = self._capture_contact_values()
        self._contact_row_count = min(needed, MAX_CONTACT_ROWS)
        self._refresh_batch_rows()
        self._restore_contact_values(captured)

    def _build_history(self) -> None:
        self.history_tab.grid_columnconfigure(0, weight=1)
        self.history_tab.grid_rowconfigure(1, weight=1)
        header = tk.Frame(self.history_tab, bg=PANEL)
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(18, 8))
        tk.Label(header, text=tr("历史记录管理"), bg=PANEL, fg=INK, font=("Microsoft YaHei UI", 14, "bold")).pack(side="left")
        tk.Label(header, text=tr("删除操作无法撤销"), bg=PANEL, fg=RED, font=("Microsoft YaHei UI", 8)).pack(side="left", padx=12)
        ttk.Button(header, text=tr("查看流水"), style="Soft.TButton", command=self.show_task_history_detail).pack(side="right", padx=(6, 0))
        ttk.Button(header, text=tr("恢复备份"), style="Soft.TButton", command=self.restore_database_backup).pack(side="right")
        ttk.Button(header, text=tr("删除选中记录"), style="Danger.TButton", command=self.delete_history_selected).pack(side="right", padx=(6, 0))
        ttk.Button(header, text=tr("清空全部记录"), style="Soft.TButton", command=self.clear_history).pack(side="right")

        card = self._card(self.history_tab, row=1, column=0, sticky="nsew", padx=20, pady=(0, 18))
        columns = ("id", "profile", "name", "email", "status", "created")
        self.history_tree = ttk.Treeview(card, columns=columns, show="headings", style="App.Treeview", selectmode="extended")
        labels = tuple(tr(label) for label in ("ID", "窗口", "姓名", "邮箱", "状态", "创建时间"))
        widths = (45, 75, 140, 240, 180, 170)
        for column, label, width in zip(columns, labels, widths):
            self.history_tree.heading(column, text=label)
            self.history_tree.column(column, width=width, minwidth=45, anchor="w")
        scroll_y = ttk.Scrollbar(card, orient="vertical", command=self.history_tree.yview)
        scroll_x = ttk.Scrollbar(card, orient="horizontal", command=self.history_tree.xview)
        self.history_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.history_tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(0, weight=1)

    def _build_window_sequence(self) -> None:
        self.window_tab.grid_columnconfigure(0, weight=1)
        self.window_tab.grid_rowconfigure(1, weight=1)
        header = tk.Frame(self.window_tab, bg=PANEL)
        header.grid(row=0, column=0, sticky="ew", padx=24, pady=(18, 8))
        tk.Label(
            header,
            text=tr("浏览器窗口顺序"),
            bg=PANEL,
            fg=INK,
            font=("Microsoft YaHei UI", 14, "bold"),
        ).pack(side="left")
        tk.Label(
            header,
            text=tr("可自由增加或删除，最多30个；任务分配后编号锁定"),
            bg=PANEL,
            fg=GREEN,
            font=("Microsoft YaHei UI", 8, "bold"),
        ).pack(side="left", padx=14)

        card = self._card(self.window_tab, row=1, column=0, sticky="nsew", padx=24, pady=(0, 18))
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(1, weight=1)

        toolbar = tk.Frame(card, bg=PAPER)
        toolbar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=18, pady=(14, 8))
        tk.Label(
            toolbar,
            text=tr("按实际需要排列窗口编号；列表顺序就是任务分配顺序"),
            bg=PAPER,
            fg=MUTED,
            font=("Microsoft YaHei UI", 9),
        ).pack(side="left")
        tk.Label(
            toolbar,
            text=tr("每行可绑定话术与发件人名字；窗口优先，未填时使用模板/全局变量"),
            bg=PAPER,
            fg=GOLD_DARK,
            font=("Microsoft YaHei UI", 8),
        ).pack(side="left", padx=(8, 0))
        ttk.Button(
            toolbar,
            text=tr("＋ 添加窗口"),
            style="Primary.TButton",
            command=self.add_window_sequence_row,
        ).pack(side="right")
        ttk.Button(
            toolbar,
            text=tr("自动获取窗口"),
            style="Soft.TButton",
            command=self.auto_fill_window_sequence,
        ).pack(side="right", padx=(0, 6))

        rows_canvas = tk.Canvas(card, bg=PAPER, highlightthickness=0, bd=0)
        rows_scroll = ttk.Scrollbar(card, orient="vertical", command=rows_canvas.yview)
        rows_canvas.configure(yscrollcommand=rows_scroll.set)
        rows_canvas.grid(row=1, column=0, sticky="nsew", padx=(18, 0))
        rows_scroll.grid(row=1, column=1, sticky="ns", padx=(0, 18))
        self.window_sequence_rows_frame = tk.Frame(rows_canvas, bg=PAPER)
        self.window_sequence_rows_frame.grid_columnconfigure(0, weight=1)
        rows_window = rows_canvas.create_window((0, 0), window=self.window_sequence_rows_frame, anchor="nw")
        self.window_sequence_rows_frame.bind(
            "<Configure>",
            lambda _event: rows_canvas.configure(scrollregion=rows_canvas.bbox("all")),
        )
        rows_canvas.bind(
            "<Configure>",
            lambda event: rows_canvas.itemconfigure(rows_window, width=event.width),
        )

        self._enable_canvas_wheel(rows_canvas, self.window_sequence_rows_frame)
        saved = list(self.settings.window_sequence)
        self.window_sequence_vars = [tk.StringVar(value=str(value)) for value in saved]
        self._render_window_sequence_rows()

        footer = tk.Frame(card, bg=PAPER)
        footer.grid(row=2, column=0, columnspan=2, sticky="ew", padx=18, pady=(10, 16))
        self.window_sequence_note_var = tk.StringVar(
            value=trf("已设置 {count} 个窗口", count=len(saved))
        )
        tk.Label(
            footer,
            textvariable=self.window_sequence_note_var,
            bg=PAPER,
            fg=MUTED,
            font=("Microsoft YaHei UI", 9),
        ).pack(side="left")
        ttk.Button(
            footer,
            text=tr("清空窗口顺序"),
            style="Soft.TButton",
            command=self.clear_window_sequence,
        ).pack(side="right", padx=(8, 0))
        ttk.Button(
            footer,
            text=tr("保存窗口顺序"),
            style="Primary.TButton",
            command=self.save_window_sequence,
        ).pack(side="right")

    def _render_window_sequence_rows(self) -> None:
        for child in self.window_sequence_rows_frame.winfo_children():
            child.destroy()
        if not self.window_sequence_vars:
            tk.Label(
                self.window_sequence_rows_frame,
                text=tr("暂无窗口编号，点击右上角“＋ 添加窗口”开始设置。"),
                bg=PAPER,
                fg=MUTED,
                font=("Microsoft YaHei UI", 9),
                pady=28,
            ).grid(row=0, column=0, sticky="ew")
            return
        self._window_template_vars: list[tk.StringVar] = []
        self._window_template_combos: list[ttk.Combobox] = []
        self._window_sender_vars: list[tk.StringVar] = []
        for index, variable in enumerate(self.window_sequence_vars):
            field_box = tk.Frame(
                self.window_sequence_rows_frame,
                bg="#FFFFFF",
                highlightthickness=1,
                highlightbackground=LINE,
            )
            field_box.grid(row=index, column=0, sticky="ew", padx=(0, 8), pady=4)
            field_box.grid_columnconfigure(1, weight=1)
            tk.Label(
                field_box,
                text=trf("顺序 {index}", index=index + 1),
                bg="#FFFFFF",
                fg=INK,
                width=10,
                anchor="w",
                font=("Microsoft YaHei UI", 9, "bold"),
            ).grid(row=0, column=0, sticky="w", padx=(12, 8), pady=8)
            ttk.Spinbox(
                field_box,
                from_=1,
                to=999,
                textvariable=variable,
                style="App.TSpinbox",
                width=16,
            ).grid(row=0, column=1, sticky="ew", padx=(0, 10), pady=7)
            binding: dict = {}
            try:
                binding = self.settings.window_bindings.get(
                    str(int(str(variable.get()).strip()))
                ) or {}
            except ValueError:
                pass
            template_var = tk.StringVar(value=binding.get("template_name", ""))
            sender_var = tk.StringVar(value=binding.get("sender_name", ""))
            self._window_template_vars.append(template_var)
            self._window_sender_vars.append(sender_var)
            template_combo = ttk.Combobox(
                field_box,
                textvariable=template_var,
                state="readonly",
                values=[
                    str(template.get("name", ""))
                    for template in self.settings.saved_templates
                    if template.get("name")
                ],
                width=16,
            )
            template_combo.grid(row=0, column=2, padx=(0, 6), pady=7)
            self._window_template_combos.append(template_combo)
            ttk.Entry(
                field_box,
                textvariable=sender_var,
                style="App.TEntry",
                width=18,
            ).grid(row=0, column=3, padx=(0, 6), pady=7)
            ttk.Button(
                field_box,
                text=tr("删除"),
                style="Danger.TButton",
                command=lambda row_index=index: self.delete_window_sequence_row(row_index),
            ).grid(row=0, column=4, padx=(0, 10), pady=6)
        self._translate_widgets()

    def add_window_sequence_row(self) -> None:
        if len(self.window_sequence_vars) >= MAX_WINDOW_SEQUENCE:
            messagebox.showwarning("已达到上限", f"窗口顺序最多只能设置 {MAX_WINDOW_SEQUENCE} 个。")
            return
        self.window_sequence_vars.append(tk.StringVar())
        self._render_window_sequence_rows()
        self.window_sequence_note_var.set(f"当前 {len(self.window_sequence_vars)} 个输入项")

    def _refresh_window_template_options(self) -> None:
        """Keep window-row template dropdowns in sync with saved templates."""
        if not hasattr(self, "_window_template_combos"):
            return
        names = [
            str(template.get("name", ""))
            for template in self.settings.saved_templates
            if template.get("name")
        ]
        for combo in self._window_template_combos:
            current = combo.get()
            combo["values"] = names
            if current and current not in names:
                combo.set("")

    def auto_fill_window_sequence(self) -> None:
        provider = create_browser_provider(self.settings)
        try:
            list_open = getattr(provider, "list_running_windows", None)
            windows = (
                list_open(verify_connection=True)
                if list_open is not None
                else provider.list_windows()
            )
        except Exception as exc:
            self._set_status(f"获取窗口失败：{exc}")
            messagebox.showerror("获取窗口失败", str(exc))
            return
        if not windows:
            messagebox.showwarning(
                tr("自动获取窗口"),
                tr("未识别到浏览器窗口，请确认浏览器应用已启动"),
            )
            return
        discovered: list[int] = []
        for number, _name in windows:
            try:
                parsed = int(number)
            except (TypeError, ValueError):
                continue
            if parsed > 0 and parsed not in discovered:
                discovered.append(parsed)
        merged: list[int] = []
        for number in self.settings.window_sequence:
            if number in discovered and number not in merged:
                merged.append(number)
        for number in sorted(discovered, reverse=True):
            if number not in merged and len(merged) < MAX_WINDOW_SEQUENCE:
                merged.append(number)
        labels = [
            f"{number} · {name}" if name else number
            for number, name in windows
        ]
        if hasattr(self, "window_picker_combo"):
            self.window_picker_combo["values"] = labels
        self.window_sequence_vars = [
            tk.StringVar(value=str(number)) for number in merged
        ]
        self._render_window_sequence_rows()
        self.window_sequence_note_var.set(
            trf("已设置 {count} 个窗口", count=len(merged))
        )
        self._set_status(
            f"已自动获取 {len(merged)} 个窗口编号，请点击“保存窗口顺序”生效"
        )

    def delete_window_sequence_row(self, row_index: int) -> None:
        if 0 <= row_index < len(self.window_sequence_vars):
            self.window_sequence_vars.pop(row_index)
            self._render_window_sequence_rows()
            self.window_sequence_note_var.set(f"当前 {len(self.window_sequence_vars)} 个输入项；保存后生效")

    def _build_settings(self) -> None:
        self.settings_sections = ttk.Notebook(self.settings_tab, style="App.TNotebook")
        self.settings_sections.pack(fill="both", expand=True, padx=18, pady=10)
        self.settings_basic_tab = tk.Frame(self.settings_sections, bg=PANEL)
        self.settings_notice_tab = tk.Frame(self.settings_sections, bg=PANEL)
        self.settings_theme_tab = tk.Frame(self.settings_sections, bg=PANEL)
        self.settings_sections.add(self.settings_basic_tab, text=tr("基础设置"))
        self.settings_sections.add(self.settings_notice_tab, text=tr("公告"))
        self.settings_sections.add(self.settings_theme_tab, text=tr("外观主题"))

        card = self._card(self.settings_basic_tab)
        card.pack(anchor="nw", fill="x", padx=18, pady=12)
        tk.Label(card, text=tr("本地与 Gmail 设置"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 14, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=22, pady=(15, 3))
        tk.Label(card, text=tr("浏览器切换已移至“任务与草稿”页面；三个本地接口地址使用内置配置。"), bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8)).grid(row=1, column=0, columnspan=2, sticky="w", padx=22, pady=(0, 8))
        self.morelogin_var = tk.StringVar(value=self.settings.morelogin_url)
        self.adspower_var = tk.StringVar(value=self.settings.adspower_url)
        self.adspower_api_key_var = tk.StringVar(value=self.settings.adspower_api_key)
        self.bitbrowser_var = tk.StringVar(value=self.settings.bitbrowser_url)
        self.sender_var = getattr(self, "sender_var", None) or tk.StringVar(
            value=self.settings.sender_name
        )
        fields = [
            (tr("AdsPower API Key（可选）"), self.adspower_api_key_var, True),
        ]
        for idx, (label, var, secret) in enumerate(fields, start=2):
            tk.Label(card, text=label, bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 9)).grid(row=idx, column=0, sticky="w", padx=22, pady=4)
            ttk.Entry(card, textvariable=var, width=62, show="*" if secret else "", style="App.TEntry").grid(row=idx, column=1, sticky="ew", padx=(8, 24), pady=4)
        tk.Label(card, text=tr("界面语言"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 9)).grid(row=3, column=0, sticky="w", padx=22, pady=4)
        self.language_var = tk.StringVar(value=LANGUAGES[self.settings.language])
        language_line = tk.Frame(card, bg=PAPER)
        language_line.grid(row=3, column=1, sticky="w", padx=(8, 24), pady=4)
        ttk.Combobox(
            language_line,
            textvariable=self.language_var,
            state="readonly",
            values=list(LANGUAGES.values()),
            width=18,
        ).pack(side="left")
        tk.Label(card, text=tr("当前版本"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 9)).grid(row=4, column=0, sticky="w", padx=22, pady=4)
        version_line = tk.Frame(card, bg=PAPER)
        version_line.grid(row=4, column=1, sticky="w", padx=(8, 24), pady=4)
        tk.Label(version_line, text=APP_VERSION, bg=RED, fg="#FFFFFF", padx=10, pady=4, font=("Segoe UI", 9, "bold")).pack(side="left")
        tk.Label(version_line, text="山水江湖主题版", bg=PAPER, fg=MUTED, font=("KaiTi", 10)).pack(side="left", padx=10)
        tk.Label(card, text=tr("管理员授权"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 9)).grid(row=5, column=0, sticky="w", padx=22, pady=4)
        self.trial_remaining_var = tk.StringVar(value=tr(remaining_text(self._trial_status)))
        tk.Label(card, textvariable=self.trial_remaining_var, bg=PAPER, fg=RED, font=("Microsoft YaHei UI", 9, "bold")).grid(row=5, column=1, sticky="w", padx=(8, 24), pady=4)
        tk.Label(card, text=tr("本机设备码"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 9)).grid(row=6, column=0, sticky="w", padx=22, pady=4)
        device_line = tk.Frame(card, bg=PAPER)
        device_line.grid(row=6, column=1, sticky="ew", padx=(8, 24), pady=4)
        self.device_code_var = tk.StringVar(value=device_code())
        ttk.Entry(device_line, textvariable=self.device_code_var, state="readonly", width=42, style="App.TEntry").pack(side="left", fill="x", expand=True)
        ttk.Button(device_line, text=tr("复制设备码"), style="Soft.TButton", command=self.copy_device_code).pack(side="left", padx=(8, 0))
        tk.Label(card, text=tr("管理员验证码"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 9)).grid(row=7, column=0, sticky="nw", padx=22, pady=4)
        code_line = tk.Frame(card, bg=PAPER)
        code_line.grid(row=7, column=1, sticky="ew", padx=(8, 24), pady=4)
        self.authorization_code_var = tk.StringVar()
        ttk.Entry(code_line, textvariable=self.authorization_code_var, style="App.TEntry").pack(side="left", fill="x", expand=True)
        ttk.Button(code_line, text=tr("立即验证"), style="Primary.TButton", command=self.verify_admin_code).pack(side="left", padx=(8, 0))
        tk.Label(
            card,
            text="授权过期后，联系人录入、本地生成和 Gmail 窗口粘贴将锁定。请把设备码发送给管理员获取验证码。",
            bg=PAPER,
            fg=MUTED,
            wraplength=690,
            justify="left",
            font=("Microsoft YaHei UI", 8),
        ).grid(row=8, column=0, columnspan=2, sticky="w", padx=22, pady=(2, 5))
        actions = tk.Frame(card, bg=PAPER)
        actions.grid(row=9, column=1, sticky="w", padx=(8, 24), pady=(5, 14))
        ttk.Button(actions, text=tr("保存设置"), style="Soft.TButton", command=self.save_settings).pack(side="left")
        ttk.Button(actions, text=tr("检查更新"), style="Soft.TButton", command=self.check_for_update).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text=tr("查看日志"), style="Soft.TButton", command=self.show_log_viewer).pack(side="left", padx=(8, 0))
        card.grid_columnconfigure(1, weight=1)

        announcement = self._card(self.settings_notice_tab)
        announcement.pack(fill="both", expand=True, padx=18, pady=12)
        tk.Label(
            announcement,
            text=tr("公 告 中 心"),
            bg=PAPER,
            fg=INK,
            font=("KaiTi", 18, "bold"),
        ).pack(pady=(18, 5))
        tk.Frame(announcement, bg=GOLD, height=2).pack(fill="x", padx=60, pady=(0, 12))
        tk.Label(
            announcement,
            text=tr("购买时长联系飞机号："),
            bg=PAPER,
            fg=MUTED,
            font=("Microsoft YaHei UI", 9),
        ).pack()
        tk.Label(
            announcement,
            text="@ls0514",
            bg=PAPER,
            fg=RED,
            font=("Segoe UI", 15, "bold"),
        ).pack(pady=(3, 10))
        if self._contact_qr_photo:
            tk.Label(
                announcement,
                image=self._contact_qr_photo,
                bg=PAPER,
                bd=0,
            ).pack(padx=18, pady=(0, 10))
        ttk.Button(
            announcement,
            text=tr("复制飞机号"),
            style="Primary.TButton",
            command=self.copy_admin_contact,
        ).pack(pady=(0, 16))
        self._build_theme_settings()

    def _build_theme_settings(self) -> None:
        card = self._card(self.settings_theme_tab)
        card.pack(anchor="nw", fill="x", padx=18, pady=12)
        tk.Label(
            card, text=tr("外观主题"), bg=PAPER, fg=INK,
            font=("Microsoft YaHei UI", 14, "bold"),
        ).grid(row=0, column=0, columnspan=4, sticky="w", padx=20, pady=(14, 4))
        tk.Label(
            card, text=tr("深浅色 + 自定义皮肤，保存后立即应用到全部页面"),
            bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8),
        ).grid(row=1, column=0, columnspan=4, sticky="w", padx=20, pady=(0, 10))

        tk.Label(
            card, text=tr("界面明暗"), bg=PAPER, fg=INK,
            font=("Microsoft YaHei UI", 9, "bold"),
        ).grid(row=2, column=0, sticky="w", padx=20, pady=6)
        mode_line = tk.Frame(card, bg=PAPER)
        mode_line.grid(row=2, column=1, sticky="w", padx=(8, 20), pady=6)
        ttk.Button(
            mode_line, text=tr("浅色"), style="Soft.TButton",
            command=lambda: self._set_theme_mode("light"),
        ).pack(side="left")

        tk.Label(
            card, text=tr("皮肤预设"), bg=PAPER, fg=INK,
            font=("Microsoft YaHei UI", 9, "bold"),
        ).grid(row=3, column=0, sticky="w", padx=20, pady=6)
        presets = [
            ("比特风浅色", {"gold": "#2563EB", "paper": "#FFFFFF", "panel": "#EFF3F8", "sidebar": "#16233F"}),
            ("紫橙新潮", {"gold": "#7C3AED", "paper": "#FFFFFF", "panel": "#F5F2FF", "sidebar": "#221A3A"}),
            ("翡翠绿", {"gold": "#0F766E", "paper": "#FFFFFF", "panel": "#EDF7F4", "sidebar": "#0B3B34"}),
        ]
        preset_line = tk.Frame(card, bg=PAPER)
        preset_line.grid(row=3, column=1, sticky="w", padx=(8, 20), pady=6)
        for name, colors in presets:
            ttk.Button(
                preset_line, text=tr(name), style="Soft.TButton",
                command=lambda label=name, palette=colors: self._apply_preset_skin(label, palette),
            ).pack(side="left", padx=(0, 8))

        tk.Label(
            card, text=tr("自定义颜色"), bg=PAPER, fg=INK,
            font=("Microsoft YaHei UI", 9, "bold"),
        ).grid(row=4, column=0, sticky="w", padx=20, pady=6)
        color_line = tk.Frame(card, bg=PAPER)
        color_line.grid(row=4, column=1, sticky="w", padx=(8, 20), pady=6)
        current = {
            "gold": GOLD, "paper": PAPER, "panel": PANEL,
            "sidebar": SIDEBAR, "ink": INK,
        }
        self._skin_color_vars: dict[str, tk.StringVar] = {}
        self._skin_swatches: dict[str, tk.Label] = {}
        for key, label in (
            ("gold", "主色"), ("paper", "卡片"), ("panel", "页面"),
            ("sidebar", "侧栏"), ("ink", "文字"),
        ):
            cell = tk.Frame(color_line, bg=PAPER)
            cell.pack(side="left", padx=(0, 10))
            var = tk.StringVar(value=current[key])
            self._skin_color_vars[key] = var
            tk.Label(cell, text=tr(label), bg=PAPER, fg=MUTED, font=("Microsoft YaHei UI", 8)).pack()
            swatch = tk.Label(
                cell, text=var.get(), bg=var.get(), fg="#FFFFFF",
                width=7, cursor="hand2", font=("Consolas", 8, "bold"),
            )
            swatch.pack(pady=2)
            self._skin_swatches[key] = swatch
            swatch.bind(
                "<Button-1>",
                lambda _event, color_key=key: self._pick_skin_color(color_key),
            )

        actions = tk.Frame(card, bg=PAPER)
        actions.grid(row=5, column=0, columnspan=4, sticky="w", padx=20, pady=(10, 16))
        ttk.Button(
            actions, text=tr("保存并应用"), style="Primary.TButton",
            command=self._apply_skin_settings,
        ).pack(side="left")
        ttk.Button(
            actions, text=tr("恢复默认"), style="Soft.TButton",
            command=self._reset_skin,
        ).pack(side="left", padx=(8, 0))
        card.grid_columnconfigure(1, weight=1)

    def _pick_skin_color(self, key: str) -> None:
        initial = self._skin_color_vars[key].get()
        _rgb, hex_color = colorchooser.askcolor(
            initial, parent=self, title="选择颜色"
        )
        if hex_color:
            self._skin_color_vars[key].set(hex_color)
            self._skin_swatches[key].configure(bg=hex_color, text=hex_color)

    def _set_theme_mode(self, mode: str) -> None:
        self.settings.theme_mode = mode
        self.settings.save()
        self._rebuild_ui()

    def _apply_preset_skin(self, name: str, palette: dict[str, str]) -> None:
        self.settings.skin_name = name
        self.settings.skin_colors = dict(palette)
        self.settings.save()
        self._rebuild_ui()

    def _apply_skin_settings(self) -> None:
        colors = {
            key: var.get().strip()
            for key, var in self._skin_color_vars.items()
            if var.get().strip()
        }
        self.settings.skin_name = "自定义"
        self.settings.skin_colors = colors
        self.settings.save()
        self._rebuild_ui()

    def _reset_skin(self) -> None:
        self.settings.skin_name = "bit-light"
        self.settings.skin_colors = {}
        self.settings.theme_mode = "light"
        self.settings.background_image = ""
        self.settings.save()
        self._rebuild_ui()

    def _build_template_editor(self) -> None:
        self.template_tab.grid_columnconfigure(0, weight=1)
        self.template_tab.grid_rowconfigure(1, weight=1)
        header = tk.Frame(self.template_tab, bg=PANEL)
        header.grid(row=0, column=0, sticky="ew", padx=24, pady=(18, 8))
        tk.Label(header, text=tr("英文邮件模板"), bg=PANEL, fg=INK, font=("Microsoft YaHei UI", 14, "bold")).pack(side="left")
        tk.Label(
            header,
            text=tr("变量已标注，可一键复制到剪贴板"),
            bg=PANEL,
            fg=GREEN,
            font=("Microsoft YaHei UI", 8, "bold"),
        ).pack(side="left", padx=14)
        self.active_template_name_label = tk.Label(
            header,
            text="",
            bg=GREEN,
            fg="#FFFFFF",
            padx=12,
            pady=4,
            font=("Microsoft YaHei UI", 9, "bold"),
        )
        self.active_template_name_label.pack(side="right")
        self._update_active_template_label()

        card_canvas = tk.Canvas(self.template_tab, bg=PANEL, highlightthickness=0, bd=0)
        card_scroll = ttk.Scrollbar(
            self.template_tab, orient="vertical", command=card_canvas.yview
        )
        card_canvas.configure(yscrollcommand=card_scroll.set)
        card_canvas.grid(row=1, column=0, sticky="nsew", padx=(24, 0), pady=(0, 18))
        card_scroll.grid(row=1, column=1, sticky="ns", pady=(0, 18))
        card_holder = tk.Frame(card_canvas, bg=PANEL)
        card_holder.grid_columnconfigure(0, weight=1)
        card_window = card_canvas.create_window((0, 0), window=card_holder, anchor="nw")

        self._enable_canvas_wheel(card_canvas, card_holder)
        card_holder.bind(
            "<Configure>",
            lambda _event: card_canvas.configure(scrollregion=card_canvas.bbox("all")),
        )
        card_canvas.bind(
            "<Configure>",
            lambda event: card_canvas.itemconfigure(card_window, width=event.width),
        )

        card = self._card(card_holder, row=0, column=0, sticky="nsew")
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(5, weight=1)

        self.variable_guide = tk.Frame(card, bg="#EAF1FF", highlightthickness=1, highlightbackground=LINE)
        self.variable_guide.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 14))
        self._render_system_variables()

        custom_guide = tk.Frame(card, bg=PAPER)
        custom_guide.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))
        custom_header = tk.Frame(custom_guide, bg=PAPER)
        custom_header.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        tk.Label(
            custom_header,
            text=tr("自定义变量 · 添加后可在模板中使用，也会显示在录入页"),
            bg=PAPER,
            fg=INK,
            font=("Microsoft YaHei UI", 9, "bold"),
        ).pack(side="left")
        ttk.Button(
            custom_header,
            text=tr("＋ 添加自定义变量"),
            style="Soft.TButton",
            command=self.add_custom_template_variable,
        ).pack(side="right")
        self.custom_variables_frame = tk.Frame(custom_guide, bg=PAPER)
        self.custom_variables_frame.grid(row=1, column=0, sticky="ew")
        self._render_custom_template_fields()
        library = tk.Frame(custom_guide, bg=PAPER)
        library.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        tk.Label(
            library,
            text=tr("模板库"),
            bg=PAPER,
            fg=INK,
            font=("Microsoft YaHei UI", 9, "bold"),
        ).pack(side="left")
        self.template_library_var = tk.StringVar()
        self.template_library_combo = ttk.Combobox(
            library,
            textvariable=self.template_library_var,
            state="readonly",
            width=26,
        )
        self.template_library_combo.pack(side="left", padx=6)
        ttk.Button(
            library,
            text=tr("保存当前为模板"),
            style="Soft.TButton",
            command=self.save_template_to_library,
        ).pack(side="left", padx=3)
        ttk.Button(
            library,
            text=tr("加载模板"),
            style="Soft.TButton",
            command=self.load_template_from_library,
        ).pack(side="left", padx=3)
        ttk.Button(
            library,
            text=tr("设为当前生效"),
            style="Primary.TButton",
            command=self.activate_selected_template,
        ).pack(side="left", padx=3)
        ttk.Button(
            library,
            text=tr("删除模板"),
            style="Soft.TButton",
            command=self.delete_template_from_library,
        ).pack(side="left", padx=3)
        tk.Label(
            custom_guide,
            text=tr("模板库仅保存，不会自动生效；未绑定窗口的任务使用“当前生效模板”"),
            bg=PAPER,
            fg=MUTED,
            font=("Microsoft YaHei UI", 8),
        ).grid(row=3, column=0, sticky="w", pady=(6, 0))
        self._refresh_template_library()

        tk.Label(card, text=tr("邮件主题"), bg=PAPER, fg=INK, font=("Microsoft YaHei UI", 9, "bold")).grid(row=2, column=0, sticky="w", padx=18, pady=(0, 5))
        self.subject_template_var = tk.StringVar(value=self.settings.subject_template)
        ttk.Entry(card, textvariable=self.subject_template_var, style="App.TEntry").grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 12))
        body_header = tk.Frame(card, bg=PAPER)
        body_header.grid(row=4, column=0, sticky="ew", padx=18, pady=(0, 5))
        tk.Label(
            body_header,
            text=tr("邮件正文"),
            bg=PAPER,
            fg=INK,
            font=("Microsoft YaHei UI", 9, "bold"),
        ).pack(side="left")
        ttk.Button(
            body_header,
            text=tr("放大编辑"),
            style="Soft.TButton",
            command=self.open_large_body_editor,
        ).pack(side="right")
        self.body_template_text = tk.Text(
            card,
            wrap="word",
            bg="#FFFFFF",
            fg=INK,
            relief="flat",
            highlightthickness=1,
            highlightbackground=LINE,
            padx=12,
            pady=12,
            font=("Segoe UI", 11),
            spacing3=4,
        )
        self.body_template_text.grid(row=5, column=0, sticky="nsew", padx=18, pady=(0, 12))
        self.body_template_text.insert("1.0", self.settings.body_template)
        tk.Label(
            card,
            text=tr("发件人名字"),
            bg=PAPER,
            fg=INK,
            font=("Microsoft YaHei UI", 9, "bold"),
        ).grid(row=6, column=0, sticky="w", padx=18, pady=(0, 5))
        self.sender_var = getattr(self, "sender_var", None) or tk.StringVar(
            value=self.settings.sender_name
        )
        ttk.Entry(
            card, textvariable=self.sender_var, style="App.TEntry",
        ).grid(row=7, column=0, sticky="ew", padx=18, pady=(0, 12))
        tk.Label(
            card,
            text=tr("发件人签名"),
            bg=PAPER,
            fg=INK,
            font=("Microsoft YaHei UI", 9, "bold"),
        ).grid(row=8, column=0, sticky="w", padx=18, pady=(0, 5))
        self.signature_text = tk.Text(
            card,
            wrap="word",
            height=3,
            bg="#FFFFFF",
            fg=INK,
            relief="flat",
            highlightthickness=1,
            highlightbackground=LINE,
            padx=12,
            pady=10,
            font=("Segoe UI", 10),
        )
        self.signature_text.grid(row=9, column=0, sticky="ew", padx=18, pady=(0, 12))
        self.signature_text.insert("1.0", self.settings.signature)
        footer = tk.Frame(card, bg=PAPER)
        footer.grid(row=10, column=0, sticky="ew", padx=18, pady=(0, 16))
        ttk.Button(footer, text=tr("恢复默认模板"), style="Soft.TButton", command=self.reset_email_template).pack(side="left")
        ttk.Button(footer, text=tr("复制主题"), style="Soft.TButton", command=self.copy_template_subject).pack(side="left", padx=(8, 0))
        ttk.Button(footer, text=tr("复制正文"), style="Soft.TButton", command=self.copy_template_body).pack(side="left", padx=(8, 0))
        ttk.Button(footer, text=tr("复制完整模板"), style="Soft.TButton", command=self.copy_full_email_template).pack(side="left", padx=(8, 0))
        ttk.Button(footer, text=tr("预览模板"), style="Soft.TButton", command=self.preview_email_template).pack(side="right", padx=(0, 8))
        ttk.Button(footer, text=tr("保存邮件模板"), style="Primary.TButton", command=self.save_email_template).pack(side="right")

    def _render_system_variables(self) -> None:
        for child in self.variable_guide.winfo_children():
            child.destroy()
        header = tk.Frame(self.variable_guide, bg="#EAF1FF")
        header.grid(row=0, column=0, columnspan=3, sticky="ew", padx=12, pady=(8, 0))
        tk.Label(
            header,
            text=tr("系统变量 · 不需要可隐藏，模板中仍可正常引用"),
            bg="#EAF1FF",
            fg=MUTED,
            font=("Microsoft YaHei UI", 8),
        ).pack(side="left")
        hidden = [
            key
            for key, _description in SYSTEM_VARIABLES
            if key in self.settings.hidden_system_variables
        ]
        if hidden:
            ttk.Button(
                header,
                text=tr("恢复变量"),
                style="Soft.TButton",
                command=self.restore_system_variable,
            ).pack(side="right", padx=(0, 6))
            self.hidden_system_var = tk.StringVar()
            ttk.Combobox(
                header,
                textvariable=self.hidden_system_var,
                state="readonly",
                values=hidden,
                width=18,
            ).pack(side="right", padx=(0, 6))
        visible = [
            item
            for item in SYSTEM_VARIABLES
            if item[0] not in self.settings.hidden_system_variables
        ]
        for column, (key, description) in enumerate(visible):
            self.variable_guide.grid_columnconfigure(column, weight=1)
            item = tk.Frame(self.variable_guide, bg="#EAF1FF")
            item.grid(row=1, column=column, sticky="nsew", padx=10, pady=10)
            token = f"{{{key}}}"
            tk.Label(
                item,
                text=token,
                bg="#EAF1FF",
                fg=GREEN,
                font=("Segoe UI", 10, "bold"),
            ).pack(anchor="w")
            tk.Label(
                item,
                text=description,
                bg="#EAF1FF",
                fg=MUTED,
                font=("Microsoft YaHei UI", 8),
            ).pack(anchor="w", pady=(2, 6))
            ttk.Button(
                item,
                text=tr("复制变量"),
                style="Soft.TButton",
                command=lambda value=token: self.copy_template_variable(value),
            ).pack(anchor="w")
            ttk.Button(
                item,
                text=tr("隐藏变量"),
                style="Soft.TButton",
                command=lambda variable_key=key: self.hide_system_variable(variable_key),
            ).pack(anchor="w", pady=(6, 0))
        self._translate_widgets()

    def hide_system_variable(self, key: str) -> None:
        if key not in self.settings.hidden_system_variables:
            self.settings.hidden_system_variables.append(key)
        try:
            self.settings.save()
        except Exception as exc:
            self._set_status(f"隐藏变量失败：{exc}")
            messagebox.showerror("隐藏变量失败", str(exc))
            return
        self._render_system_variables()
        self._rebuild_contact_rows()
        self._render_manual_custom_fields()
        self._set_status(f"已隐藏变量 {{{key}}}，需要时可从“恢复变量”加回")

    def restore_system_variable(self) -> None:
        key = getattr(self, "hidden_system_var", None)
        if key is None or not key.get():
            messagebox.showwarning("恢复变量", "请先选择要恢复的系统变量。")
            return
        selected = key.get()
        if selected in self.settings.hidden_system_variables:
            self.settings.hidden_system_variables.remove(selected)
        try:
            self.settings.save()
        except Exception as exc:
            self._set_status(f"恢复变量失败：{exc}")
            messagebox.showerror("恢复变量失败", str(exc))
            return
        self._render_system_variables()
        self._rebuild_contact_rows()
        self._render_manual_custom_fields()
        self._set_status(f"已恢复变量 {{{selected}}}")

    def _render_custom_template_fields(self) -> None:
        for child in self.custom_variables_frame.winfo_children():
            child.destroy()
        self.custom_template_vars: dict[str, tk.StringVar] = {}
        keys = list(self.settings.custom_variable_keys)
        if not keys:
            tk.Label(
                self.custom_variables_frame,
                text=tr("暂无自定义变量。点击右上角“＋ 添加自定义变量”后，把变量写入邮件主题或正文。"),
                bg=PAPER,
                fg=MUTED,
                font=("Microsoft YaHei UI", 8),
                pady=6,
            ).pack(anchor="w")
            return
        for index, key in enumerate(keys):
            item = tk.Frame(
                self.custom_variables_frame,
                bg="#FFFFFF",
                highlightthickness=1,
                highlightbackground=LINE,
            )
            item.pack(fill="x", pady=(0, 6))
            item.grid_columnconfigure(1, weight=1)
            tk.Label(
                item,
                text=f"{{{key}}}",
                bg="#FFFFFF",
                fg=GREEN,
                font=("Segoe UI", 9, "bold"),
                width=14,
                anchor="w",
            ).grid(row=0, column=0, sticky="w", padx=(8, 4), pady=6)
            value_var = tk.StringVar(value=self.settings.custom_variables.get(key, ""))
            self.custom_template_vars[key] = value_var
            ttk.Entry(item, textvariable=value_var, style="App.TEntry").grid(
                row=0, column=1, sticky="ew", padx=4, pady=6
            )
            ttk.Button(
                item,
                text=tr("复制"),
                style="Soft.TButton",
                command=lambda value=f"{{{key}}}": self.copy_template_variable(value),
            ).grid(row=0, column=2, padx=4, pady=6)
            ttk.Button(
                item,
                text=tr("移除"),
                style="Soft.TButton",
                command=lambda variable_key=key: self.remove_custom_template_variable(variable_key),
            ).grid(row=0, column=3, padx=4, pady=6)
            up_button = ttk.Button(
                item,
                text="↑",
                style="Soft.TButton",
                command=lambda variable_key=key: self.move_custom_variable(variable_key, -1),
            )
            up_button.grid(row=0, column=4, padx=2, pady=6)
            down_button = ttk.Button(
                item,
                text="↓",
                style="Soft.TButton",
                command=lambda variable_key=key: self.move_custom_variable(variable_key, 1),
            )
            down_button.grid(row=0, column=5, padx=(2, 8), pady=6)
            if index == 0:
                up_button.state(["disabled"])
            if index == len(keys) - 1:
                down_button.state(["disabled"])
        self._translate_widgets()

    def move_custom_variable(self, key: str, direction: int) -> None:
        keys = self.settings.custom_variable_keys
        if key not in keys:
            return
        index = keys.index(key)
        target = index + direction
        if target < 0 or target >= len(keys):
            return
        keys[index], keys[target] = keys[target], keys[index]
        try:
            self.settings.save()
        except Exception as exc:
            keys[index], keys[target] = keys[target], keys[index]
            self._set_status(f"变量排序保存失败：{exc}")
            messagebox.showerror("变量排序保存失败", str(exc))
            return
        self._render_custom_template_fields()
        self._rebuild_contact_rows()
        self._render_manual_custom_fields()
        self._set_status(f"变量 {key} 已{'上移' if direction < 0 else '下移'}")

    def _update_active_template_label(self) -> None:
        if not hasattr(self, "active_template_name_label"):
            return
        name = str(getattr(self.settings, "active_template_name", "") or "").strip()
        if name:
            text = trf("当前生效：{name}", name=name)
        else:
            text = tr("当前生效：自定义模板（未命名）")
        self.active_template_name_label.configure(text=text)

    def activate_selected_template(self) -> None:
        name = self.template_library_var.get()
        template = next(
            (
                item
                for item in self.settings.saved_templates
                if item.get("name") == name
            ),
            None,
        )
        if not template:
            messagebox.showwarning(
                "设为当前生效",
                "请先选择模板库中已保存的模板。",
            )
            return
        self.subject_template_var.set(template.get("subject_template", ""))
        self.body_template_text.delete("1.0", "end")
        self.body_template_text.insert("1.0", template.get("body_template", ""))
        self.sender_var.set(template.get("sender_name", self.settings.sender_name))
        self.signature_text.delete("1.0", "end")
        self.signature_text.insert("1.0", template.get("signature", ""))
        stored = template.get("custom_variables") or {}
        for key, var in self.custom_template_vars.items():
            var.set(str(stored.get(key, "")))
        self.save_email_template()
        self.settings.active_template_name = str(name)
        try:
            self.settings.save()
        except Exception:
            pass
        self._update_active_template_label()
        self._set_status(trf("模板“{name}”已设为当前生效", name=name))

    def _refresh_template_library(self) -> None:
        if not hasattr(self, "template_library_combo"):
            return
        names = [
            str(template.get("name", ""))
            for template in self.settings.saved_templates
            if template.get("name")
        ]
        self.template_library_combo["values"] = names
        current = self.template_library_var.get()
        if current not in names:
            self.template_library_var.set(names[0] if names else "")

    def save_template_to_library(self) -> None:
        from tkinter import simpledialog

        name = simpledialog.askstring("保存模板", "模板名称：", parent=self)
        if not name:
            return
        entry = {
            "name": str(name).strip(),
            "subject_template": self.subject_template_var.get().strip(),
            "body_template": self.body_template_text.get("1.0", "end-1c").strip(),
            "custom_variables": self._current_custom_variables(),
            "sender_name": self.sender_var.get().strip(),
            "signature": self.signature_text.get("1.0", "end-1c").strip(),
        }
        existing = [
            template
            for template in self.settings.saved_templates
            if template.get("name") == entry["name"]
        ]
        if existing:
            existing[0].update(entry)
        else:
            self.settings.saved_templates.append(entry)
        try:
            self.settings.save()
        except Exception as exc:
            self._set_status(f"模板保存失败：{exc}")
            messagebox.showerror("模板保存失败", str(exc))
            return
        self._refresh_template_library()
        self._refresh_window_template_options()
        self._set_status(f"模板“{entry['name']}”已保存")

    def load_template_from_library(self) -> None:
        name = self.template_library_var.get()
        template = next(
            (
                item
                for item in self.settings.saved_templates
                if item.get("name") == name
            ),
            None,
        )
        if not template:
            messagebox.showwarning("加载模板", "请先选择一个模板。")
            return
        self.subject_template_var.set(template.get("subject_template", ""))
        self.body_template_text.delete("1.0", "end")
        self.body_template_text.insert("1.0", template.get("body_template", ""))
        self.sender_var.set(template.get("sender_name", self.settings.sender_name))
        self.signature_text.delete("1.0", "end")
        self.signature_text.insert("1.0", template.get("signature", ""))
        stored = template.get("custom_variables") or {}
        for key, var in self.custom_template_vars.items():
            var.set(str(stored.get(key, "")))
        self._set_status(
            f"已加载模板“{name}”到编辑区；点击“设为当前生效”或“保存邮件模板”后生效"
        )

    def delete_template_from_library(self) -> None:
        name = self.template_library_var.get()
        if not name:
            return
        if not messagebox.askyesno(
            "删除模板",
            f"确定删除模板“{name}”吗？",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        self.settings.saved_templates = [
            template
            for template in self.settings.saved_templates
            if template.get("name") != name
        ]
        try:
            self.settings.save()
        except Exception as exc:
            self._set_status(f"删除模板失败：{exc}")
            messagebox.showerror("删除模板失败", str(exc))
            return
        self._refresh_template_library()
        self._refresh_window_template_options()
        self._set_status(f"模板“{name}”已删除")

    def add_custom_template_variable(self) -> None:
        key = next_custom_variable_key(self.settings.custom_variable_keys)
        self.settings.custom_variable_keys.append(key)
        self.settings.custom_variables[key] = ""
        try:
            self.settings.save()
        except Exception as exc:
            self.settings.custom_variable_keys.remove(key)
            self.settings.custom_variables.pop(key, None)
            self._set_status(f"添加变量失败：{exc}")
            messagebox.showerror("添加变量失败", str(exc))
            return
        self._render_custom_template_fields()
        self._rebuild_contact_rows()
        self._render_manual_custom_fields()
        self._set_status(f"已添加变量 {{{key}}}，可在模板和录入页中使用")

    def remove_custom_template_variable(self, key: str) -> None:
        if key not in self.settings.custom_variable_keys:
            return
        if not messagebox.askyesno(
            "移除自定义变量",
            f"确定移除变量 {{{key}}} 吗？模板中引用它的位置将无法预览，已生成邮件不受影响。",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        self.settings.custom_variable_keys.remove(key)
        self.settings.custom_variables.pop(key, None)
        try:
            self.settings.save()
        except Exception as exc:
            self.settings.custom_variable_keys.append(key)
            self.settings.custom_variables[key] = ""
            self._set_status(f"移除变量失败：{exc}")
            messagebox.showerror("移除变量失败", str(exc))
            return
        self._render_custom_template_fields()
        self._rebuild_contact_rows()
        self._render_manual_custom_fields()
        self._set_status(f"已移除变量 {{{key}}}")

    def _current_template_values(self) -> tuple[str, str]:
        return self.subject_template_var.get().strip(), self.body_template_text.get("1.0", "end-1c").strip()

    def open_large_body_editor(self) -> None:
        dialog = tk.Toplevel(self)
        dialog.title(tr("放大编辑邮件正文"))
        dialog.geometry("860x620")
        dialog.configure(bg=PAPER)
        text = tk.Text(
            dialog,
            wrap="word",
            bg="#FFFFFF",
            fg=INK,
            relief="flat",
            highlightthickness=1,
            highlightbackground=LINE,
            padx=16,
            pady=16,
            font=("Segoe UI", 12),
            spacing1=3,
            spacing3=6,
            undo=True,
        )
        text.pack(fill="both", expand=True, padx=14, pady=(14, 8))
        text.insert("1.0", self.body_template_text.get("1.0", "end-1c"))
        text.focus_set()

        buttons = tk.Frame(dialog, bg=PAPER)
        buttons.pack(fill="x", padx=14, pady=(0, 14))

        def save() -> None:
            self.body_template_text.delete("1.0", "end")
            self.body_template_text.insert("1.0", text.get("1.0", "end-1c"))
            dialog.destroy()
            self._set_status(tr("正文已更新，点击“保存邮件模板”后生效"))

        ttk.Button(buttons, text=tr("保存"), style="Primary.TButton", command=save).pack(side="right")
        ttk.Button(buttons, text=tr("取消"), style="Soft.TButton", command=dialog.destroy).pack(side="right", padx=(0, 8))
        dialog.bind("<Escape>", lambda _event: dialog.destroy())
        dialog.transient(self)
        dialog.grab_set()
        dialog.wait_window()

    def _current_custom_variables(self) -> dict[str, str]:
        return {key: variable.get() for key, variable in self.custom_template_vars.items()}

    def _copy_template_text(self, text: str, status: str) -> None:
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
            self.update_idletasks()
        except tk.TclError as exc:
            self._set_status(f"复制失败：{exc}")
            messagebox.showerror("复制失败", str(exc))
            return
        self._set_status(status)

    def copy_template_variable(self, token: str) -> None:
        self._copy_template_text(token, f"已复制变量 {token}")

    def copy_template_subject(self) -> None:
        subject_template, _ = self._current_template_values()
        if not subject_template:
            messagebox.showwarning("无法复制", "邮件主题为空，请先填写主题。")
            return
        self._copy_template_text(subject_template, "邮件主题模板已复制")

    def copy_template_body(self) -> None:
        _, body_template = self._current_template_values()
        if not body_template:
            messagebox.showwarning("无法复制", "邮件正文为空，请先填写正文。")
            return
        self._copy_template_text(body_template, "邮件正文模板已复制")

    def copy_full_email_template(self) -> None:
        subject_template, body_template = self._current_template_values()
        if not subject_template or not body_template:
            messagebox.showwarning("无法复制", "邮件主题和正文都需要填写。")
            return
        self._copy_template_text(f"Subject: {subject_template}\n\n{body_template}", "完整邮件模板已复制")

    def preview_email_template(self) -> None:
        from .mail_content import render_email

        subject_template, body_template = self._current_template_values()
        try:
            subject, body = render_email(
                "Alex",
                "Seattle",
                self.sender_var.get().strip() or DEFAULT_SENDER_NAME,
                subject_template,
                body_template,
                self._current_custom_variables(),
            )
        except ValueError as exc:
            messagebox.showerror("模板错误", str(exc))
            return
        messagebox.showinfo("邮件模板预览", f"Subject: {subject}\n\n{body}")

    def save_email_template(self) -> None:
        from .mail_content import render_email

        subject_template, body_template = self._current_template_values()
        try:
            render_email(
                "Alex",
                "Seattle",
                self.sender_var.get().strip() or DEFAULT_SENDER_NAME,
                subject_template,
                body_template,
                self._current_custom_variables(),
            )
        except ValueError as exc:
            messagebox.showerror("模板错误", str(exc))
            return
        self.settings.subject_template = subject_template
        self.settings.body_template = body_template
        self.settings.sender_name = self.sender_var.get().strip()
        self.settings.signature = self.signature_text.get("1.0", "end-1c").strip()
        self.settings.custom_variables = self._current_custom_variables()
        self.settings.active_template_name = ""
        missing_keys = sorted(
            {
                match.group(1)
                for match in re.finditer(
                    r"\{custom_\d+\}", subject_template + "\n" + body_template
                )
                if match.group(1) not in self.settings.custom_variable_keys
            }
        )
        for key in missing_keys:
            self.settings.custom_variable_keys.append(key)
            self.settings.custom_variables[key] = ""
        if missing_keys:
            self._render_custom_template_fields()
        try:
            self.settings.save()
        except Exception as exc:
            self._set_status(f"邮件模板保存失败：{exc}")
            messagebox.showerror("邮件模板保存失败", str(exc))
            return
        self.workflow = Workflow(self.db, self.settings)
        self._rebuild_contact_rows()
        self._render_manual_custom_fields()
        suffix = (
            "，已自动补回变量："
            + "、".join(f"{{{key}}}" for key in missing_keys)
            if missing_keys
            else ""
        )
        self._set_status("邮件模板已保存" + suffix)
        self._update_active_template_label()

    def reset_email_template(self) -> None:
        if not messagebox.askyesno(
            "恢复默认模板",
            "确定恢复默认英文邮件模板吗？",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        self.subject_template_var.set(DEFAULT_SUBJECT_TEMPLATE)
        self.body_template_text.delete("1.0", "end")
        self.body_template_text.insert("1.0", DEFAULT_BODY_TEMPLATE)
        self.save_email_template()

    def add_contacts(self) -> None:
        if not self._require_authorization():
            return
        errors: list[str] = []
        filled = 0
        valid_rows: list[tuple[str, str, str, int, dict[str, str]]] = []
        hidden = set(self.settings.hidden_system_variables)
        for row_number, (name_var, location_var, email_var, custom_vars) in enumerate(self.contact_rows, start=1):
            name = name_var.get().strip()
            location = location_var.get().strip()
            email = email_var.get().strip()
            if not name and not location and not email:
                continue
            filled += 1
            if (
                (not name and "first_name" not in hidden)
                or (not location and "location" not in hidden)
                or "@" not in email
            ):
                errors.append(f"第 {row_number} 行")
                continue
            custom_values = {
                key: value_var.get().strip()
                for key, value_var in custom_vars.items()
                if value_var.get().strip()
            }
            valid_rows.append((name, location, email, row_number - 1, custom_values))
        if filled == 0:
            messagebox.showwarning("缺少资料", "请至少填写一组名字、地区和邮箱地址。")
            return
        added = 0
        if valid_rows:
            try:
                task_ids = self.db.add_local_tasks(
                    [
                        (name, location, email)
                        for name, location, email, _index, _custom in valid_rows
                    ],
                    [custom for _name, _location, _email, _index, custom in valid_rows],
                )
                for task_id, (_name, _location, _email, index, _custom) in zip(task_ids, valid_rows):
                    if task_id is None:
                        errors.append(f"第 {index + 1} 行（任务已存在）")
                        continue
                    added += 1
                    for value_var in self._row_string_vars(self.contact_rows[index]):
                        value_var.set("")
            except Exception as exc:
                errors.append(f"批量导入失败（{exc}）")
        self.refresh()
        self._set_status(f"已添加 {added} 条联系人" + (f"，{len(errors)} 条未导入" if errors else ""))
        if errors:
            messagebox.showwarning("部分资料未导入", "以下输入需要检查：\n" + "、".join(errors))
        if added:
            self._select_page(self.queue_tab)

    @staticmethod
    def _row_string_vars(
        row_vars: tuple[tk.StringVar, tk.StringVar, tk.StringVar, dict[str, tk.StringVar]],
    ) -> tuple[tk.StringVar, ...]:
        return tuple(row_vars[:3]) + tuple(row_vars[3].values())

    def clear_contact_rows(self) -> None:
        filled = sum(
            1
            for row_vars in self.contact_rows
            if any(
                value_var.get().strip()
                for value_var in self._row_string_vars(row_vars)
            )
        )
        if filled == 0:
            self._set_status("当前没有已填写的联系人资料")
            return
        if not messagebox.askyesno(
            "确认清空输入",
            f"确定清空全部 {filled} 行已填写的联系人资料吗？此操作无法撤销。",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        for row_vars in self.contact_rows:
            for value_var in self._row_string_vars(row_vars):
                value_var.set("")
        self._set_status("已清空联系人输入")

    def import_contacts_xlsx(self) -> None:
        from tkinter import filedialog

        path = filedialog.askopenfilename(
            title="选择联系人 XLSX",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return
        try:
            entries = parse_contacts_xlsx(path)
        except Exception as exc:
            self._set_status(f"导入 XLSX 失败：{exc}")
            messagebox.showerror("导入 XLSX 失败", str(exc))
            return
        if not entries:
            messagebox.showwarning("导入 XLSX", "文件中没有可导入的数据（模板示例行不计入）。")
            return
        if len(entries) > MAX_CONTACT_ROWS:
            messagebox.showwarning(
                "导入 XLSX",
                f"文件共 {len(entries)} 行，只导入前 {MAX_CONTACT_ROWS} 行。",
            )
            entries = entries[:MAX_CONTACT_ROWS]
        entries = self._confirm_xlsx_entries(entries)
        if entries is None:
            return
        self._ensure_contact_rows(len(entries))

        filled = 0
        for index, entry in enumerate(entries):
            name_var, location_var, email_var, custom_vars = self.contact_rows[index]
            name_var.set(entry["name"])
            location_var.set(entry["location"])
            email_var.set(entry["email"])
            for key, var in custom_vars.items():
                var.set(entry.get(key, entry.get(key.replace("custom_", "变量"), "")))
            filled += 1
        self._set_status(f"XLSX 已导入 {filled} 行（模板示例行已跳过），请核对后加入队列")

    def _confirm_xlsx_entries(self, entries: list[dict[str, str]]) -> list[dict[str, str]] | None:
        try:
            existing_rows = self.db.list_tasks()
        except Exception:
            existing_rows = []
        existing_emails = {
            str(row["recipient_email"]).casefold() for row in existing_rows
        }
        seen: set[str] = set()
        flagged: list[tuple[dict[str, str], str]] = []
        for entry in entries:
            problems: list[str] = []
            email = str(entry.get("email", "")).strip()
            hidden = set(self.settings.hidden_system_variables)
            if not entry.get("name") and "first_name" not in hidden:
                problems.append("缺名字")
            if not entry.get("location") and "location" not in hidden:
                problems.append("缺地区")
            if "@" not in email:
                problems.append("邮箱无效")
            elif email.casefold() in seen:
                problems.append("文件内重复")
            elif email.casefold() in existing_emails:
                problems.append("已存在")
            seen.add(email.casefold())
            flagged.append((entry, "、".join(problems)))

        result: dict[str, object] = {"entries": None}
        dialog = tk.Toplevel(self)
        dialog.title("XLSX 导入预览")
        dialog.geometry("760x520")
        dialog.configure(bg=PAPER)
        problem_count = sum(1 for _entry, problems in flagged if problems)
        tk.Label(
            dialog,
            text=f"共 {len(entries)} 行，其中 {problem_count} 行需要留意（重复/缺字段/已存在）。"
            "确认后填入录入页，不会直接入库。",
            bg=PAPER,
            fg=RED if problem_count else GREEN,
            font=("Microsoft YaHei UI", 9, "bold"),
            wraplength=720,
            justify="left",
        ).pack(anchor="w", padx=14, pady=(14, 6))

        frame = tk.Frame(dialog, bg=PAPER)
        frame.pack(fill="both", expand=True, padx=10, pady=6)
        tree = ttk.Treeview(
            frame,
            columns=("no", "name", "location", "email", "notice"),
            show="headings",
            style="App.Treeview",
        )
        for column, label, width in (
            ("no", "行号", 50),
            ("name", "名字", 120),
            ("location", "地区", 140),
            ("email", "邮箱", 240),
            ("notice", "提示", 160),
        ):
            tree.heading(column, text=label)
            tree.column(column, width=width, minwidth=40, anchor="w")
        scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        for index, (entry, problems) in enumerate(flagged, start=1):
            tree.insert(
                "",
                "end",
                values=(
                    index,
                    entry.get("name", ""),
                    entry.get("location", ""),
                    entry.get("email", ""),
                    problems,
                ),
            )

        buttons = tk.Frame(dialog, bg=PAPER)
        buttons.pack(fill="x", padx=14, pady=(4, 14))

        def confirm() -> None:
            if problem_count:
                choice = messagebox.askyesnocancel(
                    "仍有问题行",
                    "是：跳过问题行导入\n否：全部导入\n取消：不导入",
                    icon=messagebox.WARNING,
                    default=messagebox.NO,
                )
                if choice is None:
                    return
                result["entries"] = (
                    [entry for entry, problems in flagged if not problems]
                    if choice
                    else entries
                )
            else:
                result["entries"] = entries
            dialog.destroy()

        def cancel() -> None:
            result["entries"] = None
            dialog.destroy()

        ttk.Button(buttons, text="确认导入", style="Primary.TButton", command=confirm).pack(side="right")
        ttk.Button(buttons, text="取消", style="Soft.TButton", command=cancel).pack(side="right", padx=(0, 8))
        dialog.bind("<Escape>", lambda _event: cancel())
        dialog.grab_set()
        dialog.wait_window()
        return result["entries"]

    def download_xlsx_template(self) -> None:
        from tkinter import filedialog

        path = filedialog.asksaveasfilename(
            title="保存联系人 XLSX 模板",
            defaultextension=".xlsx",
            initialfile="联系人导入模板.xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return
        try:
            import openpyxl

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "联系人模板"
            headers = ["名字", "地区", "邮箱"] + [
                f"变量{index}" for index in range(1, 6)
            ]
            sheet.append(headers)
            sheet.append(
                ["张三", "Seattle", "zhangsan@example.com", "酒店推荐", "美食", "安全", "", ""]
            )
            for column, width in zip(
                "ABCDEFGH", (14, 14, 30, 14, 14, 14, 10, 10)
            ):
                sheet.column_dimensions[column].width = width
            workbook.save(path)
            workbook.close()
        except Exception as exc:
            self._set_status(f"模板保存失败：{exc}")
            messagebox.showerror("模板保存失败", str(exc))
            return
        self._set_status(f"XLSX 模板已保存：{path}")
        messagebox.showinfo(
            "XLSX 模板",
            f"模板已保存到：\n{path}\n\n"
            "第二行是模板示例，导入时会自动跳过不计入；"
            "请从第三行开始填写，最多支持 100 行。",
        )

    def _add_contact_row(self, row_index: int) -> int:
        name_var, location_var, email_var, custom_vars = self.contact_rows[row_index]
        name = name_var.get().strip()
        location = location_var.get().strip()
        email = email_var.get().strip()
        if not name and not location and not email:
            raise ValueError(f"第 {row_index + 1} 行尚未填写")
        hidden = set(self.settings.hidden_system_variables)
        if not name and "first_name" not in hidden:
            raise ValueError(f"第 {row_index + 1} 行名字为空")
        if not location and "location" not in hidden:
            raise ValueError(f"第 {row_index + 1} 行地区为空")
        if "@" not in email:
            raise ValueError(f"第 {row_index + 1} 行邮箱地址无效")
        custom_values = {
            key: value_var.get().strip()
            for key, value_var in custom_vars.items()
            if value_var.get().strip()
        }
        task_id = self.db.add_local_task(name, location, email, custom_values)
        for value_var in self._row_string_vars(self.contact_rows[row_index]):
            value_var.set("")
        return task_id

    def generate_contact_row(self, row_index: int) -> None:
        if not self._require_authorization():
            return
        try:
            task_id = self._add_contact_row(row_index)
        except Exception as exc:
            self._show_generation_error("无法生成邮件", exc)
            return
        self._bind_profiles_at_generation([task_id])
        self.refresh()
        self._select_page(self.queue_tab)
        self._background_task(
            lambda cancel_event: self.workflow.generate_local(task_id, cancel_event=cancel_event),
            f"正在本地生成第 {row_index + 1} 行邮件",
            on_success=lambda: self._select_task_by_id(task_id),
            on_error=lambda exc: self._show_generation_error("邮件生成失败", exc),
        )

    def generate_all_contacts(self) -> None:
        if not self._require_authorization():
            return
        task_ids: list[tuple[int, int]] = []
        skipped: list[str] = []
        valid_rows: list[tuple[int, str, str, str, dict[str, str]]] = []
        hidden = set(self.settings.hidden_system_variables)
        for row_index, (name_var, location_var, email_var, custom_vars) in enumerate(self.contact_rows):
            name = name_var.get().strip()
            location = location_var.get().strip()
            email = email_var.get().strip()
            if not name and not location and not email:
                continue
            if not name and "first_name" not in hidden:
                skipped.append(f"第 {row_index + 1} 行名字为空")
                continue
            if not location and "location" not in hidden:
                skipped.append(f"第 {row_index + 1} 行地区为空")
                continue
            if "@" not in email:
                skipped.append(f"第 {row_index + 1} 行邮箱地址无效")
                continue
            custom_values = {
                key: value_var.get().strip()
                for key, value_var in custom_vars.items()
                if value_var.get().strip()
            }
            valid_rows.append((row_index, name, location, email, custom_values))
        if valid_rows:
            created = self.db.add_local_tasks(
                [
                    (name, location, email)
                    for _row, name, location, email, _custom in valid_rows
                ],
                [custom for _row, _name, _location, _email, custom in valid_rows],
            )
            for task_id, (row_index, _name, _location, _email, _custom) in zip(created, valid_rows):
                if task_id is None:
                    skipped.append(f"第 {row_index + 1} 行任务已存在")
                    continue
                task_ids.append((row_index + 1, task_id))
                for value_var in self._row_string_vars(self.contact_rows[row_index]):
                    value_var.set("")
        if not task_ids:
            messagebox.showwarning("没有可处理资料", "请填写至少一组有效的名字、地区和邮箱地址。")
            return
        self._bind_profiles_at_generation([task_id for _row_number, task_id in task_ids])
        self.refresh()
        self._select_page(self.queue_tab)

        cancel_event, operation_serial = self.operations.begin()
        self._set_operation_busy(True)

        def coordinator() -> None:
            succeeded = 0
            runtime_failures: list[str] = []
            total = len(task_ids)
            started_at = time.monotonic()
            self.after(0, lambda: self._set_status(
                f"正在本地生成 {total} 封邮件",
                busy=True,
            ))
            for index, (row_number, task_id) in enumerate(task_ids, start=1):
                if cancel_event.is_set():
                    break
                try:
                    self.workflow.generate_local(task_id, cancel_event)
                    succeeded += 1
                except OperationCancelledError:
                    break
                except Exception as exc:
                    runtime_failures.append(f"第 {row_number} 行：{exc}")
                    logger.exception("第 %s 行生成失败", row_number)
                    try:
                        self.db.update_task(task_id, last_error=str(exc))
                    except Exception:
                        pass
                elapsed = time.monotonic() - started_at
                average = elapsed / index
                remaining = average * (total - index)
                self.after(
                    0,
                    lambda done=index, eta=remaining: self._set_status(
                        f"正在本地生成 {done}/{total} 封，"
                        f"预计还需 {self._format_eta(eta)}",
                        busy=True,
                    ),
                )

            if not self.operations.is_current(cancel_event, operation_serial):
                return

            failures = list(skipped) + runtime_failures

            def finish() -> None:
                elapsed = int(time.monotonic() - started_at)
                cancelled = total - succeeded - len(runtime_failures)
                self._set_operation_busy(False)
                self.refresh()
                self._set_status(
                    f"已生成 {succeeded} 封，跳过 {len(failures)} 条，"
                    f"取消 {cancelled} 条，用时 {elapsed} 秒"
                )
                if failures:
                    details = "\n".join(failures[:MAX_CONCURRENT_TASKS])
                    hint = self._generation_fix_hint_text(
                        runtime_failures[0] if runtime_failures else ""
                    )
                    messagebox.showwarning(
                        "批量生成完成",
                        f"成功：{succeeded}\n跳过：{len(failures)}\n"
                        f"取消：{cancelled}\n用时：{elapsed} 秒\n\n{details}\n\n"
                        f"解决方案：{hint}",
                    )
                else:
                    self._set_status(f"已成功生成 {succeeded} 封邮件，用时 {elapsed} 秒")

            self.after(0, finish)
            self.operations.finish(cancel_event, operation_serial)

        threading.Thread(target=coordinator, daemon=True).start()

    def _current_window_sequence(self) -> list[int]:
        return normalize_window_sequence([var.get() for var in self.window_sequence_vars])

    def _bind_profiles_at_generation(self, task_ids: list[int]) -> dict[int, int]:
        """Assign windows at generation time so window templates render first."""
        if not task_ids:
            return {}
        try:
            sequence = self._current_window_sequence()
        except ValueError as exc:
            self._set_status(f"窗口顺序无效：{exc}")
            return {}
        if not sequence:
            self._set_status("未设置窗口顺序，生成将使用当前生效模板")
            return {}
        rows = self.db.get_tasks(task_ids)
        existing = [int(row["profile_no"] or 0) for row in rows]
        pending = self.db.pending_counts_by_window()
        resolved = resolve_task_windows_balanced(existing, sequence, pending)
        assigned: dict[int, int] = {}
        for task_id, profile_no in zip(task_ids, resolved):
            if profile_no is None or int(profile_no) <= 0:
                continue
            try:
                self.db.update_task(task_id, profile_no=int(profile_no))
            except ValueError:
                continue
            assigned[int(task_id)] = int(profile_no)
        return assigned

    def _regenerate_pending_for_windows(self, profile_numbers: list[int]) -> None:
        """Regenerate unsent tasks after their window template binding changes."""
        rows = self.db.pending_tasks_by_profiles(profile_numbers)
        if not rows:
            return
        succeeded = 0
        failures: list[str] = []
        for row in rows:
            try:
                self.workflow.generate_local(int(row["id"]))
                succeeded += 1
            except Exception as exc:
                failures.append(f"任务 {row['id']}：{exc}")
        self.refresh()
        if failures:
            self._set_status(
                f"已按窗口模板重新生成 {succeeded} 封，失败 {len(failures)} 封"
            )
            messagebox.showwarning(
                "模板同步未完成",
                "部分任务重新生成失败：\n"
                + "\n".join(failures[:MAX_CONCURRENT_TASKS]),
            )
        else:
            self._set_status(
                f"已按窗口模板重新生成 {succeeded} 封待发送邮件"
            )

    def save_window_sequence(self) -> None:
        if not self._require_authorization():
            return
        try:
            sequence = self._current_window_sequence()
        except ValueError as exc:
            messagebox.showerror("窗口顺序无效", str(exc))
            return
        old_bindings = dict(self.settings.window_bindings or {})
        bindings: dict[str, dict] = dict(old_bindings)
        template_vars = getattr(self, "_window_template_vars", [])
        sender_vars = getattr(self, "_window_sender_vars", [])
        for index, number in enumerate(sequence):
            template_name = (
                template_vars[index].get() if index < len(template_vars) else ""
            )
            sender = (
                sender_vars[index].get().strip()
                if index < len(sender_vars)
                else ""
            )
            bindings[str(number)] = {
                "template_name": template_name,
                "sender_name": sender,
                "locked": bool(
                    (old_bindings.get(str(number)) or {}).get("locked")
                ),
            }
        try:
            self.settings.save_window_sequence(sequence)
            self.settings.save_window_bindings(bindings)
        except Exception as exc:
            self._set_status(f"窗口顺序保存失败：{exc}")
            messagebox.showerror("窗口顺序保存失败", str(exc))
            return
        self.window_sequence_vars = [tk.StringVar(value=str(value)) for value in sequence]
        self._render_window_sequence_rows()
        self.window_sequence_note_var.set(f"已设置 {len(sequence)} 个窗口")
        detail = "、".join(str(number) for number in sequence)
        self._set_status(f"窗口顺序已保存：{detail}")
        changed = [
            number
            for number in sequence
            if old_bindings.get(str(number)) != bindings.get(str(number))
        ]
        if changed:
            self._regenerate_pending_for_windows(changed)

    def clear_window_sequence(self) -> None:
        if not self.window_sequence_vars:
            self._set_status("当前没有可清空的窗口编号")
            return
        if not messagebox.askyesno(
            "确认清空窗口顺序",
            "确定清空当前全部窗口编号吗？已锁定到任务的编号不会改变。",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        self.window_sequence_vars.clear()
        self._render_window_sequence_rows()
        try:
            self.settings.save_window_sequence([])
        except Exception as exc:
            self._set_status(f"清空窗口顺序失败：{exc}")
            messagebox.showerror("清空窗口顺序失败", str(exc))
            return
        self.window_sequence_note_var.set("已设置 0 个窗口")
        self._set_status("窗口顺序已清空；已经锁定到任务的编号不会改变")

    def _selected_task_ids_in_display_order(self) -> list[int]:
        selected = set(self.tree.selection())
        return [
            int(self.tree.item(item, "values")[0])
            for item in self.tree.get_children()
            if item in selected
        ]

    def _draft_profiles(
        self, task_ids: list[int]
    ) -> tuple[list[tuple[int, int]], list[str], list[int], bool]:
        """Resolve each task's browser window without persisting anything."""
        sequence = self._current_window_sequence()
        rows = self.db.get_tasks(task_ids)
        existing_profiles = [int(row["profile_no"] or 0) for row in rows]
        resolved_profiles = resolve_task_windows(existing_profiles, sequence)
        runnable_tasks: list[tuple[int, int]] = []
        failures: list[str] = []
        sequence_used = False
        for index, (task_id, profile_no, row) in enumerate(
            zip(task_ids, resolved_profiles, rows)
        ):
            if row["status"] in {"sent", "replied"}:
                failures.append(f"任务 {task_id}：已标记为已发送/已回复")
                continue
            if not row["subject"] or not row["body"]:
                failures.append(f"任务 {task_id}：尚未生成完整邮件预览")
                continue
            if profile_no is None or profile_no <= 0:
                failures.append(f"任务 {task_id}：尚未选择发送窗口")
                continue
            if existing_profiles[index] <= 0:
                sequence_used = True
            runnable_tasks.append((task_id, profile_no))
        return runnable_tasks, failures, sequence, sequence_used

    def open_selected_drafts(self, wait_send: bool = False) -> None:
        if not self._require_authorization():
            return
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("未选择任务", "请按住 Ctrl 选择需要填写 Gmail 的任务。")
            return
        if len(selected) > MAX_CONCURRENT_TASKS:
            messagebox.showwarning("选择过多", f"每次最多同时处理 {MAX_CONCURRENT_TASKS} 条任务。")
            return
        task_ids = self._selected_task_ids_in_display_order()
        try:
            runnable_tasks, initial_failures, sequence, sequence_used = self._draft_profiles(task_ids)
        except ValueError as exc:
            messagebox.showerror("无法按窗口顺序填写", str(exc))
            return
        runnable_count = len(runnable_tasks)
        if runnable_count == 0:
            messagebox.showwarning("没有可执行任务", "\n".join(initial_failures[:MAX_CONCURRENT_TASKS]))
            return
        estimate_seconds = runnable_count * 8

        # Follow the Window Order page: same-window tasks stay grouped and
        # windows run in the order configured there.
        rank = {int(profile_no): index for index, profile_no in enumerate(sequence)}
        runnable_tasks.sort(
            key=lambda pair: (
                rank.get(int(pair[1]), len(sequence) + int(pair[1])),
                pair[0],
            )
        )

        for task_id, profile_no in runnable_tasks:
            row = self.db.get_task(task_id)
            if row is not None and int(row["profile_locked"] or 0) != 1:
                self.db.lock_task_profile(task_id, profile_no)
        try:
            self.settings.save_window_sequence(sequence)
        except Exception as exc:
            self._set_status(f"窗口顺序保存失败：{exc}")
            messagebox.showerror("窗口顺序保存失败", str(exc))
            return
        self.window_sequence_vars = [tk.StringVar(value=str(value)) for value in sequence]
        self._render_window_sequence_rows()
        self.window_sequence_note_var.set(f"已设置 {len(sequence)} 个窗口")

        profile_locks = {
            profile_no: threading.Lock()
            for _task_id, profile_no in runnable_tasks
        }

        cancel_event, operation_serial = self.operations.begin()
        self._set_operation_busy(True)

        def process_task(task_id: int, profile_no: int) -> None:
            # Tasks aimed at the same browser must never edit one compose box
            # concurrently; different windows run in parallel.
            with profile_locks[profile_no]:
                if wait_send:
                    self.workflow.open_draft_wait_send(
                        task_id, cancel_event=cancel_event
                    )
                else:
                    self.workflow.open_draft(task_id, cancel_event=cancel_event)

        def coordinator() -> None:
            succeeded = 0
            runtime_failures: list[str] = []
            failed_task_ids: list[int] = []
            processed = 0
            profile_failures: dict[int, int] = {}
            started_at = time.monotonic()
            mode = "优先按任务表编号，空缺使用窗口顺序" if sequence_used else "按任务表已填编号"
            action_text = "填写并自动发送" if wait_send else "填写"
            self.after(0, lambda: self._set_status(
                f"正在{mode}{action_text} {runnable_count} 封邮件，"
                f"预计等待约 {self._format_eta(estimate_seconds)}",
                busy=True,
            ))
            with self._worker_pool as executor:
                futures = {}
                for index, (task_id, profile_no) in enumerate(runnable_tasks):
                    if cancel_event.is_set():
                        break
                    if profile_failures.get(profile_no, 0) >= 3:
                        runtime_failures.append(
                            f"窗口 {profile_no} / 任务 {task_id}："
                            "该窗口已连续失败 3 次，自动跳过剩余任务"
                        )
                        failed_task_ids.append(task_id)
                        continue
                    future = executor.submit(process_task, task_id, profile_no)
                    futures[future] = (task_id, profile_no)
                for future in as_completed(futures):
                    processed += 1
                    task_id, profile_no = futures[future]
                    try:
                        future.result()
                        succeeded += 1
                    except OperationCancelledError:
                        pass
                    except Exception as exc:
                        runtime_failures.append(f"窗口 {profile_no} / 任务 {task_id}：{exc}")
                        failed_task_ids.append(task_id)
                        profile_failures[profile_no] = (
                            profile_failures.get(profile_no, 0) + 1
                        )
                        logger.exception(
                            "窗口 %s / 任务 %s 填写失败",
                            profile_no,
                            task_id,
                        )
                    elapsed = time.monotonic() - started_at
                    average = elapsed / processed
                    remaining = average * (runnable_count - processed)
                    self.after(
                        0,
                        lambda done=processed, eta=remaining: self._set_status(
                            f"正在{action_text} {done}/{runnable_count}，"
                            f"预计还需 {self._format_eta(eta)}",
                            busy=True,
                        ),
                    )

            if not self.operations.is_current(cancel_event, operation_serial):
                return

            failures = list(initial_failures) + runtime_failures

            def finish() -> None:
                elapsed = int(time.monotonic() - started_at)
                cancelled = runnable_count - succeeded - len(runtime_failures)
                self._set_operation_busy(False)
                self.refresh()
                self._last_failed_task_ids = set(failed_task_ids)
                self._update_retry_button()
                done_text = "已发送" if wait_send else "已填写"
                self._set_status(
                    f"{done_text} {succeeded} 封，跳过 {len(failures)} 条，"
                    f"取消 {cancelled} 条，用时 {elapsed} 秒"
                )
                if failures:
                    summary = (
                        f"成功：{succeeded}\n跳过：{len(failures)}\n"
                        f"取消：{cancelled}\n用时：{elapsed} 秒"
                    )
                    messagebox.showwarning(
                        "批量发送完成" if wait_send else "批量填写完成",
                        summary + "\n\n"
                        + "\n".join(failures[:MAX_CONCURRENT_TASKS]),
                    )
                else:
                    profile_count = len({profile_no for _task_id, profile_no in runnable_tasks})
                    self._set_status(
                        (
                            f"已在 {profile_count} 个窗口中发送 {succeeded} 封邮件，"
                            f"用时 {elapsed} 秒"
                        )
                        if wait_send
                        else (
                            f"已在 {profile_count} 个窗口中填写 {succeeded} 封邮件草稿，"
                            f"用时 {elapsed} 秒"
                        )
                    )

            self.after(0, finish)
            self.operations.finish(cancel_event, operation_serial)

        threading.Thread(target=coordinator, daemon=True).start()

    def retry_failed_drafts(self) -> None:
        if not self._last_failed_task_ids:
            self._set_status("当前没有可重试的失败任务")
            return
        rows = self.db.get_tasks(list(self._last_failed_task_ids))
        existing = [
            int(row["id"])
            for row in rows
            if row["status"] not in {"sent", "replied"}
        ]
        if not existing:
            self._last_failed_task_ids = set()
            self._update_retry_button()
            self._set_status("失败任务已不存在或已发送，无需重试")
            return
        self._select_task_ids(existing)
        self.open_selected_drafts()

    def _select_task_ids(self, task_ids: list[int]) -> None:
        self.tree.selection_remove(*self.tree.selection())
        wanted = set(task_ids)
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if values and int(values[0]) in wanted:
                self.tree.selection_add(item)
                self.tree.see(item)
                wanted.discard(int(values[0]))

    def _update_retry_button(self) -> None:
        if not hasattr(self, "retry_failed_button"):
            return
        if self._last_failed_task_ids:
            self.retry_failed_button.state(["!disabled"])
            self.retry_failed_button.configure(
                text=tr("重试失败草稿") + f" ({len(self._last_failed_task_ids)})"
            )
        else:
            self.retry_failed_button.state(["disabled"])
            self.retry_failed_button.configure(text=tr("重试失败草稿"))

    def select_all_queue_tasks(self) -> None:
        items = self.tree.get_children()
        if not items:
            self._set_status(tr("当前任务列表中没有邮件"))
            return
        self.tree.selection_set(items)
        self.tree.focus(items[0])
        self.tree.see(items[0])
        self.show_preview()
        self._set_status(trf("已选择全部 {count} 条邮件", count=len(items)))

    def generate_ungenerated_tasks(self) -> None:
        if not self._require_authorization():
            return
        try:
            rows = self.db.list_tasks()
        except Exception as exc:
            self._set_status(f"读取任务失败：{exc}")
            messagebox.showerror("读取任务失败", str(exc))
            return
        task_ids = [
            int(row["id"])
            for row in rows
            if row["status"] in {"pending", "new", "needs_review"}
        ]
        if not task_ids:
            self._set_status("当前没有未生成的任务")
            messagebox.showinfo("生成未生成任务", "当前所有任务都已生成邮件内容。")
            return
        if len(task_ids) > MAX_CONCURRENT_TASKS:
            self._set_status(
                f"共 {len(task_ids)} 条未生成，本次先生成前 {MAX_CONCURRENT_TASKS} 条，"
                "可再次点击继续生成"
            )
        self._select_task_ids(task_ids[:MAX_CONCURRENT_TASKS])
        self.generate_selected_tasks()

    def show_daily_stats(self) -> None:
        try:
            stats = self.db.daily_stats()
        except Exception as exc:
            self._set_status(f"统计失败：{exc}")
            messagebox.showerror("统计失败", str(exc))
            return
        lines = [
            f"今日生成：{stats['generated_today']} 封",
            f"今日填写草稿：{stats['drafted_today']} 封",
            f"今日已发送：{stats['sent_today']} 封",
            f"累计失败：{stats['failed_total']} 条",
        ]
        by_window = stats["sent_by_window"]
        if by_window:
            lines.append("\n按窗口发送：")
            lines.extend(f"窗口 {window}：{count} 封" for window, count in by_window.items())
        messagebox.showinfo("今日统计", "\n".join(lines))

    def show_window_status(self) -> None:
        try:
            status_rows = self.db.window_status()
        except Exception as exc:
            self._set_status(f"窗口状态读取失败：{exc}")
            messagebox.showerror("窗口状态读取失败", str(exc))
            return
        if not status_rows:
            messagebox.showinfo(tr("窗口状态"), tr("当前没有绑定窗口的任务"))
            return

        dialog = tk.Toplevel(self)
        dialog.title(tr("窗口状态"))
        dialog.geometry("640x420")
        dialog.configure(bg=PAPER)
        tk.Label(
            dialog,
            text=tr("每个窗口的进度：已填草稿 / 待发送 / 已发送 / 失败"),
            bg=PAPER,
            fg=INK,
            font=("Microsoft YaHei UI", 10, "bold"),
        ).pack(anchor="w", padx=14, pady=(14, 6))
        frame = tk.Frame(dialog, bg=PAPER)
        frame.pack(fill="both", expand=True, padx=10, pady=6)
        tree = ttk.Treeview(
            frame,
            columns=("window", "drafted", "pending", "sent", "failed"),
            show="headings",
            style="App.Treeview",
        )
        labels = (
            (tr("窗口"), "window", 90),
            (tr("已填草稿"), "drafted", 110),
            (tr("待发送"), "pending", 110),
            (tr("已发送"), "sent", 110),
            (tr("失败"), "failed", 90),
        )
        for text, key, width in labels:
            tree.heading(key, text=text)
            tree.column(key, width=width, minwidth=60, anchor="center")
        scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        for row in status_rows:
            tree.insert(
                "",
                "end",
                values=(
                    row["window"],
                    row["drafted"],
                    row["pending"],
                    row["sent"],
                    row["failed"],
                ),
            )
        ttk.Button(
            dialog,
            text=tr("关闭"),
            style="Soft.TButton",
            command=dialog.destroy,
        ).pack(pady=(4, 14))

    def _on_task_selection(self) -> None:
        count = len(self.tree.selection())
        self.selection_count_var.set(trf("已选择 {count} 条", count=count))
        self.show_preview()

    def _set_task_filter(self, value: str) -> None:
        self.task_filter = value
        self._render_task_rows(self._all_rows)
        labels = {
            "all": tr("全部任务"),
            "active": tr("待处理任务"),
            "new": tr("未生成"),
            "ready": tr("待确认"),
            "drafted": tr("Gmail 草稿"),
        }
        self._set_status(
            trf("当前筛选：{value}", value=labels.get(value, value))
        )

    def _schedule_filter_refresh(self) -> None:
        if getattr(self, "_filter_job", None):
            self.after_cancel(self._filter_job)
        self._filter_job = self.after(150, lambda: self._render_task_rows(self._all_rows))

    def cancel_current_operation(self) -> None:
        cancel_event, serial = self.operations.begin()
        self.operations.finish(cancel_event, serial)
        self._set_operation_busy(False)
        self._set_status("已停止当前任务；可以立即开始新的操作")

    def _select_all_shortcut(self, _event=None) -> str:
        self.select_all_queue_tasks()
        return "break"

    def generate_selected_tasks(self) -> None:
        if not self._require_authorization():
            return
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("未选择邮件", "请先选择邮件，或点击“一键全选”。")
            return
        if len(selected) > MAX_CONCURRENT_TASKS:
            messagebox.showwarning(
                "选择过多",
                f"每次最多同时生成 {MAX_CONCURRENT_TASKS} 封邮件，请减少选择数量。",
            )
            return
        task_ids = [int(self.tree.item(item, "values")[0]) for item in selected]
        self._bind_profiles_at_generation(task_ids)

        cancel_event, operation_serial = self.operations.begin()
        self._set_operation_busy(True)

        def coordinator() -> None:
            succeeded = 0
            runtime_failures: list[str] = []
            total = len(task_ids)
            started_at = time.monotonic()
            self.after(
                0,
                lambda: self._set_status(
                    f"正在本地生成选中的 {total} 封邮件",
                    busy=True,
                ),
            )
            for index, task_id in enumerate(task_ids, start=1):
                if cancel_event.is_set():
                    break
                try:
                    self.workflow.generate_local(task_id, cancel_event)
                    succeeded += 1
                except OperationCancelledError:
                    break
                except Exception as exc:
                    runtime_failures.append(f"任务 {task_id}：{exc}")
                    logger.exception("任务 %s 生成失败", task_id)
                    try:
                        self.db.update_task(task_id, last_error=str(exc))
                    except Exception:
                        pass
                elapsed = time.monotonic() - started_at
                average = elapsed / index
                remaining = average * (total - index)
                self.after(
                    0,
                    lambda done=index, eta=remaining: self._set_status(
                        f"正在本地生成 {done}/{total} 封，"
                        f"预计还需 {self._format_eta(eta)}",
                        busy=True,
                    ),
                )

            if not self.operations.is_current(cancel_event, operation_serial):
                return

            failures = runtime_failures

            def finish() -> None:
                elapsed = int(time.monotonic() - started_at)
                cancelled = total - succeeded - len(runtime_failures)
                self._set_operation_busy(False)
                self.refresh()
                self._set_status(
                    f"已生成 {succeeded} 封，跳过 {len(failures)} 条，"
                    f"取消 {cancelled} 条，用时 {elapsed} 秒"
                )
                if failures:
                    summary = (
                        f"成功：{succeeded}\n跳过：{len(failures)}\n"
                        f"取消：{cancelled}\n用时：{elapsed} 秒"
                    )
                    hint = self._generation_fix_hint_text(
                        runtime_failures[0] if runtime_failures else ""
                    )
                    messagebox.showwarning(
                        "所选邮件生成完成",
                        summary + "\n\n"
                        + "\n".join(failures[:MAX_CONCURRENT_TASKS])
                        + f"\n\n解决方案：{hint}",
                    )
                else:
                    self._set_status(f"已成功生成全部 {succeeded} 封邮件，用时 {elapsed} 秒")

            self.after(0, finish)
            self.operations.finish(cancel_event, operation_serial)

        threading.Thread(target=coordinator, daemon=True).start()

    def selected_id(self, warn: bool = True) -> int | None:
        selected = self.tree.selection()
        if not selected:
            if warn:
                messagebox.showwarning("未选择任务", "请先选择一条任务。")
            return None
        return int(self.tree.item(selected[0], "values")[0])

    def mark_selected_sent(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("未选择任务", "请先选择已经在 Gmail 中发送的任务。")
            return
        task_ids = [int(self.tree.item(item, "values")[0]) for item in selected]
        if not messagebox.askyesno(
            "标记已发送",
            f"确定把选中的 {len(task_ids)} 条任务标记为已发送吗？\n\n"
            "标记后任务会从工作台移除，可在历史记录中查看。",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        try:
            updated = self.db.mark_sent(task_ids)
        except Exception as exc:
            self._set_status(f"标记失败：{exc}")
            messagebox.showerror("标记失败", str(exc))
            return
        self.refresh()
        self.preview.configure(state="normal")
        self.preview.delete("1.0", "end")
        self.preview.configure(state="disabled")
        self._set_status(f"已标记 {updated} 条任务为已发送，已移出工作台")

    def unmark_selected_sent(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("未选择任务", "请先选择已标记为已发送的任务。")
            return
        task_ids = [int(self.tree.item(item, "values")[0]) for item in selected]
        rows = self.db.get_tasks(task_ids)
        sent_ids = [
            int(row["id"])
            for row in rows
            if row["status"] in {"sent", "replied"}
        ]
        if not sent_ids:
            self._set_status("所选任务不是已发送状态，无需撤销")
            return
        if not messagebox.askyesno(
            "撤销标记",
            f"确定把选中的 {len(sent_ids)} 条任务恢复为 Gmail 草稿吗？",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        try:
            updated = self.db.unmark_sent(sent_ids)
        except Exception as exc:
            self._set_status(f"撤销失败：{exc}")
            messagebox.showerror("撤销失败", str(exc))
            return
        self.refresh()
        self._set_status(f"已撤销 {updated} 条任务的已发送标记")

    def delete_selected_queue_tasks(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("未选择资料", "请先在任务表中选择要删除的人物或草稿记录。")
            return
        task_ids = [int(self.tree.item(item, "values")[0]) for item in selected]
        rows = [self.db.get_task(task_id) for task_id in task_ids]
        draft_count = sum(1 for row in rows if row is not None and row["status"] == "drafted")
        if len(rows) == 1 and rows[0] is not None:
            row = rows[0]
            name = row["name_override"] or " ".join(
                value for value in (row["first_name"], row["last_name"]) if value
            ) or "未命名联系人"
            summary = f"{name}\n{row['recipient_email']}"
        else:
            summary = f"共 {len(task_ids)} 条人物/草稿记录"
        draft_note = f"\n\n其中包含 {draft_count} 条 Gmail 草稿记录。" if draft_count else ""
        warning = (
            f"确定从牛马邮箱中删除以下资料吗？\n\n{summary}{draft_note}"
            "\n\n此操作无法撤销。Gmail 中已经打开的真实草稿不会被删除。"
        )
        if not messagebox.askyesno(
            "确认删除人物/草稿",
            warning,
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        try:
            self.db.delete_tasks(task_ids)
        except Exception as exc:
            self._set_status(f"删除失败：{exc}")
            messagebox.showerror("删除失败", str(exc))
            return
        self.refresh()
        self.preview.configure(state="normal")
        self.preview.delete("1.0", "end")
        self.preview.configure(state="disabled")
        self.manual_name_var.set("")
        self.manual_location_var.set("")
        if hasattr(self, "manual_sender_var"):
            self.manual_sender_var.set("")
        self.profile_assign_note_var.set("请选择任务")
        self._set_status(f"已删除 {len(task_ids)} 条人物/草稿记录")

    @staticmethod
    def _status_text(row) -> str:
        return tr(STATUS_LABELS.get(row["status"], row["status"]))

    def _select_task_by_id(self, task_id: int) -> bool:
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if values and int(values[0]) == task_id:
                self.tree.selection_set(item)
                self.tree.focus(item)
                self.tree.see(item)
                return True
        return False

    def _on_window_picker_selected(self, _event=None) -> None:
        value = self.window_picker_var.get()
        if not value:
            return
        number = str(value).split(" ", 1)[0]
        try:
            self.profile_assign_var.set(int(number))
        except ValueError:
            pass

    def _fetch_windows_async(self, quiet: bool = False) -> None:
        serial = getattr(self, "_picker_fetch_serial", 0) + 1
        self._picker_fetch_serial = serial
        provider = create_browser_provider(self.settings)

        def runner() -> None:
            try:
                windows = provider.list_windows()
            except Exception as exc:
                logger.warning("识别窗口失败：%s", exc, exc_info=True)
                def fail() -> None:
                    if self._picker_fetch_serial != serial:
                        return
                    if not quiet:
                        self._set_status(f"识别窗口失败：{exc}")
                        messagebox.showerror("识别窗口失败", str(exc))

                self.after(0, fail)
                return

            def finish() -> None:
                if self._picker_fetch_serial != serial:
                    return
                if not windows:
                    if not quiet:
                        self._set_status("未识别到浏览器窗口，请确认浏览器应用已启动")
                        messagebox.showwarning(
                            tr("自动识别窗口"),
                            tr("未识别到浏览器窗口，请确认浏览器应用已启动"),
                        )
                    return
                labels = [
                    f"{number} · {name}" if name else number
                    for number, name in windows
                ]
                self.window_picker_combo["values"] = labels
                suffix = "" if quiet else "，可直接在下拉中选择"
                self._set_status(f"已识别 {len(windows)} 个浏览器窗口{suffix}")

            self.after(0, finish)

        threading.Thread(target=runner, daemon=True).start()

    def auto_populate_picker_quiet(self) -> None:
        if self.window_picker_combo["values"]:
            return
        self._fetch_windows_async(quiet=True)

    def refresh_window_picker(self) -> None:
        self._fetch_windows_async(quiet=False)

    def assign_profile_to_selected(self) -> None:
        if not self._require_authorization():
            return
        task_id = self.selected_id()
        if task_id is None:
            return
        row = self.db.get_task(task_id)
        if int(row["profile_locked"] or 0) == 1:
            messagebox.showwarning(
                "窗口编号已锁定",
                f"任务 {task_id} 已锁定到窗口 {row['profile_no']}。只有删除任务才能解除。",
            )
            return
        try:
            profile_no = int(self.profile_assign_var.get())
        except (TypeError, ValueError, tk.TclError):
            messagebox.showerror("窗口编号无效", "请输入有效的浏览器窗口编号。")
            return
        if profile_no <= 0:
            messagebox.showerror("窗口编号无效", "浏览器窗口编号必须大于 0。")
            return
        if not messagebox.askyesno(
            "确认锁定浏览器窗口",
            f"任务 {task_id} 将锁定到窗口 {profile_no}，锁定后只能删除任务解除。确定吗？",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        try:
            self.db.lock_task_profile(task_id, profile_no)
        except Exception as exc:
            messagebox.showerror("无法保存窗口", str(exc))
            return
        try:
            self.workflow.generate_local(task_id)
        except Exception as exc:
            self._set_status(
                f"任务 {task_id} 已锁定窗口 {profile_no}，但按窗口模板重新生成失败"
            )
            messagebox.showwarning(
                "内容未刷新",
                f"任务 {task_id} 已锁定窗口 {profile_no}，"
                f"但按窗口模板重新生成失败：\n{exc}",
            )
        else:
            self._set_status(
                f"任务 {task_id} 已锁定窗口 {profile_no}，并按窗口模板重新生成"
            )
        self.refresh()
        self.show_preview()

    def _handle_task_table_click(self, event) -> None:
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        column = self.tree.identify_column(event.x)
        if column not in {"#2", "#3", "#4"}:
            return
        if not self._require_authorization():
            return
        item = self.tree.identify_row(event.y)
        if not item:
            return
        self.tree.selection_set(item)
        if column == "#2":
            self.after_idle(lambda selected_item=item: self._start_inline_profile_edit(selected_item))
        else:
            self.after_idle(
                lambda selected_item=item, selected_column=column: self._start_inline_cell_edit(
                    selected_item, selected_column
                )
            )

    def _start_inline_profile_edit(self, item: str) -> None:
        values = self.tree.item(item, "values")
        if not values:
            return
        bbox = self.tree.bbox(item, "#2")
        if not bbox:
            return
        task_id = int(values[0])
        row = self.db.get_task(task_id)
        if int(row["profile_locked"] or 0) == 1:
            self._set_status(
                f"任务 {task_id} 已锁定到窗口 {row['profile_no']}，只有删除任务才能解除"
            )
            return
        profile_var = tk.StringVar(value=str(row["profile_no"] if row["profile_no"] > 0 else 1))
        x, y, width, height = bbox
        editor = ttk.Spinbox(
            self.tree,
            from_=1,
            to=999,
            textvariable=profile_var,
            style="App.TSpinbox",
        )
        editor.place(x=x, y=y, width=width, height=height)
        self._inline_editor = editor
        finished = False

        def finish(save: bool) -> None:
            nonlocal finished
            if finished:
                return
            finished = True
            raw_value = profile_var.get().strip()
            editor.destroy()
            if getattr(self, "_inline_editor", None) is editor:
                self._inline_editor = None
            if not save:
                return
            try:
                profile_no = int(raw_value)
                if profile_no <= 0:
                    raise ValueError
            except (TypeError, ValueError, tk.TclError):
                self._set_status("窗口编号无效：请输入大于 0 的数字")
                return
            current_row = self.db.get_task(task_id)
            current_profile = int(current_row["profile_no"] or 0) if current_row else 0
            if current_profile == profile_no:
                return
            if current_row is not None and int(current_row["profile_locked"] or 0) == 1:
                self._set_status(
                    f"任务 {task_id} 已锁定到窗口 {current_profile}，只有删除任务才能解除"
                )
                return
            if not messagebox.askyesno(
                "确认锁定浏览器窗口",
                f"任务 {task_id} 将锁定到窗口 {profile_no}，锁定后只能删除任务解除。确定吗？",
                icon=messagebox.WARNING,
                default=messagebox.NO,
            ):
                return
            try:
                self.db.lock_task_profile(task_id, profile_no)
            except Exception as exc:
                self._set_status(f"无法保存窗口：{exc}")
                return
            try:
                self.workflow.generate_local(task_id)
            except Exception as exc:
                self._set_status(
                    f"任务 {task_id} 已锁定窗口 {profile_no}，但按窗口模板重新生成失败"
                )
                messagebox.showwarning(
                    "内容未刷新",
                    f"任务 {task_id} 已锁定窗口 {profile_no}，"
                    f"但按窗口模板重新生成失败：\n{exc}",
                )
            else:
                self._set_status(
                    f"任务 {task_id} 已锁定窗口 {profile_no}，并按窗口模板重新生成"
                )
            self.refresh()
            self._select_task_by_id(task_id)
            self.show_preview()

        editor.bind("<Return>", lambda _event: finish(True))
        editor.bind("<Escape>", lambda _event: finish(False))
        editor.bind("<FocusOut>", lambda _event: finish(True))
        editor.focus_set()
        editor.selection_range(0, "end")

    def _start_inline_cell_edit(self, item: str, column: str) -> None:
        values = self.tree.item(item, "values")
        if not values:
            return
        bbox = self.tree.bbox(item, column)
        if not bbox:
            return
        task_id = int(values[0])
        row = self.db.get_task(task_id)
        current_name = row["name_override"] or " ".join(
            value for value in (row["first_name"], row["last_name"]) if value
        )
        current_location = row["location_override"] or row["location"]
        field_is_name = column == "#3"
        edit_var = tk.StringVar(value=current_name if field_is_name else current_location)
        x, y, width, height = bbox
        editor = ttk.Entry(self.tree, textvariable=edit_var, style="App.TEntry")
        editor.place(x=x, y=y, width=width, height=height)
        self._inline_editor = editor
        finished = False

        def finish(save: bool) -> None:
            nonlocal finished
            if finished:
                return
            finished = True
            value = edit_var.get().strip()
            editor.destroy()
            if getattr(self, "_inline_editor", None) is editor:
                self._inline_editor = None
            if not save:
                return
            current_value = current_name if field_is_name else current_location
            if value == current_value:
                return
            field_label = "姓名" if field_is_name else "地点"
            if not messagebox.askyesno(
                "确认保存并重新生成",
                f"将把{field_label}改为“{value}”并重新生成邮件预览，确定吗？",
                icon=messagebox.WARNING,
                default=messagebox.NO,
            ):
                return
            name = value if field_is_name else current_name
            location = current_location if field_is_name else value
            try:
                self.workflow.apply_manual_profile(task_id, name, location)
            except Exception as exc:
                self._set_status(f"资料保存失败：{exc}")
                return
            self.refresh()
            self._select_task_by_id(task_id)
            self.show_preview()
            self._set_status(f"任务 {task_id} 已更新，右侧邮件预览已重新生成")

        editor.bind("<Return>", lambda _event: finish(True))
        editor.bind("<Escape>", lambda _event: finish(False))
        editor.bind("<FocusOut>", lambda _event: finish(True))
        editor.focus_set()
        editor.selection_range(0, "end")

    def save_manual_profile(self) -> None:
        if not self._require_authorization():
            return
        task_id = self.selected_id()
        if task_id is None:
            return
        custom_values = {
            key: var.get().strip()
            for key, var in self.manual_custom_vars.items()
            if var.get().strip()
        }
        try:
            self.db.update_task(
                task_id,
                custom_variables=json.dumps(
                    custom_values, ensure_ascii=False, sort_keys=True
                ),
            )
            self.workflow.apply_manual_profile(
                task_id,
                self.manual_name_var.get(),
                self.manual_location_var.get(),
                self.manual_sender_var.get(),
            )
        except Exception as exc:
            messagebox.showerror("无法保存资料", str(exc))
            return
        self.refresh()
        self._select_task_by_id(task_id)
        self.show_preview()
        self._set_status("手动资料已保存，右侧邮件预览已重新生成")

    @staticmethod
    def _task_row_values(row) -> tuple:
        name = row["name_override"] or " ".join(
            value for value in (row["first_name"], row["last_name"]) if value
        )
        if int(row["profile_locked"] or 0) == 1:
            profile_text = f"{row['profile_no']}（已锁定）"
        elif row["profile_no"] > 0:
            profile_text = f"{row['profile_no']}（点击修改）"
        else:
            profile_text = "点击选择"
        location = row["location_override"] or row["location"]
        return (
            row["id"], profile_text, name or "点击填写",
            location or "点击填写", row["recipient_email"], App._status_text(row),
        )

    def _filtered_rows(self, rows):
        query = self.task_search_var.get().strip().casefold() if hasattr(self, "task_search_var") else ""
        result = []
        for row in rows:
            status = row["status"]
            if self.task_filter == "active":
                if status in {"sent", "replied"}:
                    continue
            elif self.task_filter != "all" and not (
                status == self.task_filter
                or self.task_filter == "new"
                and status in {"pending", "new", "needs_review"}
            ):
                continue
            values = self._task_row_values(row)
            if query and query not in " ".join(str(value) for value in values).casefold():
                continue
            result.append((row, values))
        return result

    def _render_task_rows(self, rows) -> None:
        if not hasattr(self, "tree"):
            return
        selected_ids = {
            int(self.tree.item(item, "values")[0])
            for item in self.tree.selection()
            if self.tree.item(item, "values")
        }
        desired = self._filtered_rows(rows)
        desired_iids = {str(int(row["id"])) for row, _values in desired}
        for item in self.tree.get_children():
            if item not in desired_iids:
                self.tree.delete(item)
        for index, (row, values) in enumerate(desired):
            iid = str(int(row["id"]))
            snapshot = tuple(values)
            if self.tree.exists(iid):
                if self._row_snapshots.get(int(row["id"])) != snapshot:
                    self.tree.item(iid, values=values)
                self.tree.move(iid, "", index)
            else:
                self.tree.insert("", index, iid=iid, values=values)
            self._row_snapshots[int(row["id"])] = snapshot
        visible_selected = [str(task_id) for task_id in selected_ids if self.tree.exists(str(task_id))]
        if visible_selected:
            self.tree.selection_set(visible_selected)
        self.selection_count_var.set(
            trf("已选择 {count} 条", count=len(self.tree.selection()))
        )

    def refresh(self) -> None:
        try:
            rows = self.db.list_tasks()
        except Exception as exc:
            self._set_status(f"刷新失败：{exc}")
            messagebox.showerror("刷新失败", str(exc))
            return
        self._all_rows = rows
        self._render_task_rows(rows)
        counts = Counter(row["status"] for row in rows)
        self.stat_vars["all"].set(str(len(rows)))
        self.stat_vars["new"].set(
            str(counts["pending"] + counts["new"] + counts["needs_review"])
        )
        self.stat_vars["ready"].set(
            str(counts["generated"] + counts["ready"])
        )
        self.stat_vars["drafted"].set(str(counts["drafted"]))
        if hasattr(self, "ungenerated_hint_var"):
            ungenerated = (
                counts["pending"] + counts["new"] + counts["needs_review"]
            )
            self.ungenerated_hint_var.set(
                f"有 {ungenerated} 条任务未生成邮件内容" if ungenerated else ""
            )
        if hasattr(self, "history_tab") and self.notebook.nametowidget(self.notebook.select()) is self.history_tab:
            self.refresh_history(rows)

    def refresh_history(self, rows=None) -> None:
        if not hasattr(self, "history_tree"):
            return
        rows = rows if rows is not None else self.db.list_tasks()
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        for row in rows:
            name = row["name_override"] or " ".join(value for value in (row["first_name"], row["last_name"]) if value)
            self.history_tree.insert("", "end", values=(
                row["id"], row["profile_no"] if row["profile_no"] > 0 else "未选择", name, row["recipient_email"],
                self._status_text(row), row["created_at"],
            ))

    def show_preview(self) -> None:
        task_id = self.selected_id(warn=False)
        if task_id is None:
            return
        row = self.db.get_task(task_id)
        effective_name = row["name_override"] or " ".join(
            value for value in (row["first_name"], row["last_name"]) if value
        )
        self.manual_name_var.set(effective_name)
        self.manual_location_var.set(row["location_override"] or row["location"])
        if hasattr(self, "manual_sender_var"):
            self.manual_sender_var.set(row["sender_name_override"] or "")
        if hasattr(self, "manual_custom_vars"):
            task_custom: dict[str, str] = {}
            raw_custom = row["custom_variables"] or ""
            if raw_custom:
                try:
                    parsed_custom = json.loads(raw_custom)
                except ValueError:
                    parsed_custom = {}
                if isinstance(parsed_custom, dict):
                    task_custom = {
                        str(key): str(value)
                        for key, value in parsed_custom.items()
                    }
            for key, var in self.manual_custom_vars.items():
                var.set(
                    task_custom.get(
                        key, self.settings.custom_variables.get(key, "")
                    )
                )
        preview_subject, preview_body = row["subject"], row["body"]
        if preview_subject and preview_body:
            text = f"Subject: {preview_subject}\n\n{preview_body}"
        else:
            text = "尚未生成邮件。请检查名字和地区，然后点击“本地生成所选”。"
        self.preview.configure(state="normal")
        self.preview.delete("1.0", "end")
        self.preview.insert("1.0", text)
        self.preview.configure(state="disabled")
        if row["profile_no"] > 0:
            self.profile_assign_var.set(row["profile_no"])
            if int(row["profile_locked"] or 0) == 1:
                self.profile_assign_note_var.set(
                    trf(
                        "已锁定：窗口 {window}（删除任务才可解除）",
                        window=row["profile_no"],
                    )
                )
            else:
                self.profile_assign_note_var.set(
                    trf("当前：窗口 {window}", window=row["profile_no"])
                )
        else:
            self.profile_assign_note_var.set(tr("尚未选择窗口"))
        if int(row["profile_locked"] or 0) == 1:
            self.profile_assign_button.state(["disabled"])
        else:
            self.profile_assign_button.state(["!disabled"])

    def delete_history_selected(self) -> None:
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("未选择记录", "请先选择要删除的历史记录。")
            return
        task_ids = [int(self.history_tree.item(item, "values")[0]) for item in selected]
        description = f" ID {task_ids[0]} 的" if len(task_ids) == 1 else f"选中的 {len(task_ids)} 条"
        if not messagebox.askyesno(
            "确认删除",
            f"确定永久删除{description}记录吗？此操作无法撤销。",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        try:
            self.db.delete_tasks(task_ids)
        except Exception as exc:
            self._set_status(f"删除失败：{exc}")
            messagebox.showerror("删除失败", str(exc))
            return
        self.refresh()
        self._set_status(f"已删除 {len(task_ids)} 条历史记录")

    def clear_history(self) -> None:
        try:
            count = len(self.db.list_tasks())
        except Exception as exc:
            self._set_status(f"读取记录失败：{exc}")
            messagebox.showerror("读取记录失败", str(exc))
            return
        if count == 0:
            self._set_status("当前没有历史记录")
            return
        if not messagebox.askyesno(
            "清空全部记录",
            f"确定永久删除全部 {count} 条记录吗？此操作无法撤销。",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        try:
            removed = self.db.clear_tasks()
        except Exception as exc:
            self._set_status(f"清空失败：{exc}")
            messagebox.showerror("清空失败", str(exc))
            return
        self.refresh()
        self._set_status(f"已清空 {removed} 条历史记录")

    def show_task_history_detail(self) -> None:
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("未选择记录", "请先选择一条记录。")
            return
        task_id = int(self.history_tree.item(selected[0], "values")[0])
        row = self.db.get_task(task_id)
        if row is None:
            return
        messagebox.showinfo(
            f"任务流水 #{task_id}",
            f"收件人：{row['recipient_email']}\n"
            f"姓名：{row['name_override'] or row['first_name'] or '未填写'}\n"
            f"地点：{row['location_override'] or row['location']}\n"
            f"窗口：{row['profile_no'] if row['profile_no'] > 0 else '未选择'}\n"
            f"状态：{self._status_text(row)}\n"
            f"创建：{row['created_at']}\n"
            f"生成：{row['generated_at'] or '未生成'}\n"
            f"填写草稿：{row['drafted_at'] or '未填写'}\n"
            f"发送：{row['sent_at'] or '未发送'}\n"
            f"尝试次数：{row['attempts'] or 0}\n"
            f"失败原因：{row['last_error'] or '无'}\n\n"
            f"主题：{row['subject'] or '未生成'}",
        )

    def restore_database_backup(self) -> None:
        backups = self.db.list_backups()
        if not backups:
            messagebox.showinfo("恢复备份", "当前没有可用备份。")
            return
        if not messagebox.askyesno(
            "恢复备份",
            "恢复备份会覆盖当前任务数据，确定继续吗？",
            icon=messagebox.WARNING,
            default=messagebox.NO,
        ):
            return
        from tkinter import filedialog

        path = filedialog.askopenfilename(
            title="选择备份文件",
            initialdir=str(self.db.path.parent / "backups"),
            filetypes=[("数据库备份", "ophelia_*.db")],
        )
        if not path:
            return
        try:
            if not self.db.restore_backup(path):
                messagebox.showerror("恢复失败", "备份文件不存在或无法恢复。")
                return
        except Exception as exc:
            self._set_status(f"恢复失败：{exc}")
            messagebox.showerror("恢复失败", str(exc))
            return
        self.refresh()
        self._set_status(f"已从备份恢复：{path}")
        messagebox.showinfo(
            "恢复备份",
            "已从备份恢复。为避免数据不一致，请重启软件后继续操作。",
        )

    @staticmethod
    def _generation_fix_hint_text(text: str) -> str:
        if "请填写联系人名字" in text or "名字为空" in text:
            return "请在录入页填写「名字」；若名字已隐藏，可在模板页恢复该变量。"
        if "城市" in text or "地区" in text:
            return "请填写有效的城市或地区（至少 2 个字符，不能是纯数字）。"
        if "占位符" in text:
            return "模板里有无效变量，请到「邮件模板」页只使用标注的变量（{first_name}、{location}、{sender_name}、{custom_1} 等）。"
        if "为空" in text:
            return "请检查模板的主题/正文是否为空，并确认已保存并设为当前生效模板。"
        if "Row" in text and "get" in text:
            return "数据读取异常，请升级到最新版本后重新生成。"
        return "请检查：名字/地区是否填写、模板是否设为当前生效、变量是否有效；仍失败请把这条提示发给管理员。"

    @classmethod
    def _generation_fix_hint(cls, exc: Exception) -> str:
        return cls._generation_fix_hint_text(str(exc))

    def _show_generation_error(self, title: str, exc: Exception) -> None:
        messagebox.showerror(
            title,
            f"{exc}\n\n解决方案：{self._generation_fix_hint(exc)}",
        )

    def _background_task(self, action, label: str, on_success=None, on_error=None) -> None:
        cancel_event, operation_serial = self.operations.begin()
        self._set_operation_busy(True)

        def runner() -> None:
            self.after(0, lambda: self._set_status(label, busy=True))
            try:
                action(cancel_event)
            except OperationCancelledError:
                return
            except Exception as exc:
                logger.exception("后台任务失败：%s", label)
                if self.operations.is_current(cancel_event, operation_serial):
                    if on_error is not None:
                        self.after(0, lambda exc=exc: on_error(exc))
                    else:
                        self.after(
                            0,
                            lambda exc=exc: messagebox.showerror(
                                "操作失败",
                                f"{label.replace('正在', '').strip()}失败：{exc}",
                            ),
                        )
                    self.after(
                        0,
                        lambda exc=exc: self._set_status(
                            f"{label.replace('正在', '').strip()}失败：{exc}"
                        ),
                    )
            else:
                if self.operations.is_current(cancel_event, operation_serial):
                    self.after(
                        0,
                        lambda: self._set_status(
                            f"{label.replace('正在', '').strip()}完成"
                        ),
                    )
                    if on_success is not None:
                        self.after(0, on_success)
            finally:
                was_current = self.operations.is_current(cancel_event, operation_serial)
                if was_current:
                    self.after(0, self.refresh)
                    self.after(0, self.show_preview)
                self.operations.finish(cancel_event, operation_serial)
                if was_current:
                    self.after(0, lambda: self._set_operation_busy(False))
        self._worker_pool.submit(runner)

    @staticmethod
    def _format_eta(seconds: float) -> str:
        seconds = max(0, int(round(seconds)))
        if seconds < 60:
            return f"约 {seconds} 秒"
        minutes, remainder = divmod(seconds, 60)
        if remainder:
            return f"约 {minutes} 分 {remainder} 秒"
        return f"约 {minutes} 分钟"

    def _set_status(self, text: str, busy: bool = False) -> None:
        self.status_var._niuma_orig_value = text
        self.status_var.set(tr(text))
        state = tr("处理中…") if busy else tr("本地模式")
        self.connection_var.set(f"{tr(remaining_text(self._trial_status))} · {state}")

    def _set_operation_busy(self, busy: bool) -> None:
        self._operation_busy = busy
        if getattr(self, "_busy_watchdog_job", None):
            self.after_cancel(self._busy_watchdog_job)
            self._busy_watchdog_job = None
        if busy:
            self._busy_watchdog_job = self.after(
                45 * 60 * 1000, self._force_busy_reset
            )
        work_state = "disabled" if busy else "!disabled"
        stop_state = "!disabled" if busy else "disabled"
        for button in getattr(self, "_operation_buttons", []):
            button.state([work_state])
        if hasattr(self, "cancel_operation_button"):
            self.cancel_operation_button.state([stop_state])

    def _force_busy_reset(self) -> None:
        self._busy_watchdog_job = None
        logger.warning("检测到任务超过 45 分钟未完成，已自动恢复按钮状态")
        self._set_operation_busy(False)
        self._set_status("任务执行异常已自动恢复，可重新操作")

    def _apply_authorization_state(self) -> None:
        remaining = remaining_text(self._trial_status)
        remaining = tr(remaining)
        mode_text = (
            tr("本地模式")
            if self._trial_status.active
            else tr("功能已锁定")
        )
        if getattr(self._trial_status, "suspicious", False):
            mode_text = tr("授权异常，请核对")
        self.connection_var.set(
            f"{remaining} · "
            f"{mode_text}"
        )
        if hasattr(self, "trial_remaining_var"):
            trial_text = remaining
            if getattr(self._trial_status, "suspicious", False):
                trial_text += "\n" + tr(
                    "检测到授权状态异常，请核对（已暂停自动写回）"
                )
            self.trial_remaining_var.set(trial_text)
        if hasattr(self, "profile_assign_button"):
            if self._trial_status.active:
                self.profile_assign_button.state(["!disabled"])
            else:
                self.profile_assign_button.state(["disabled"])
        if hasattr(self, "lock_banner"):
            if self._trial_status.active:
                self.lock_banner.pack_forget()
                self.lock_banner_var.set("")
            else:
                self.lock_banner_var.set(
                    tr(
                        "授权已过期：联系人录入、本地生成和 Gmail 窗口粘贴已锁定，"
                        "请到“系统设置”输入管理员验证码。"
                    )
                )
                self.lock_banner.pack(fill="x", side="top", before=self.shell_body)

    def _show_authorization_required(self) -> None:
        if hasattr(self, "settings_tab"):
            self.notebook.select(self.settings_tab)
            self._sync_nav()
        self._set_status("功能已锁定，请先输入管理员验证码")

    def _require_authorization(self) -> bool:
        self._trial_status = check_trial()
        self._apply_authorization_state()
        if self._trial_status.active:
            return True
        self._show_authorization_required()
        return False

    def copy_device_code(self) -> None:
        try:
            self.clipboard_clear()
            self.clipboard_append(self.device_code_var.get())
            self.update_idletasks()
        except tk.TclError as exc:
            self._set_status(f"复制失败：{exc}")
            messagebox.showerror("复制失败", str(exc))
            return
        self._set_status("设备码已复制，请发送给管理员")

    def copy_admin_contact(self) -> None:
        try:
            self.clipboard_clear()
            self.clipboard_append("@ls0514")
            self.update_idletasks()
        except tk.TclError as exc:
            self._set_status(f"复制失败：{exc}")
            messagebox.showerror("复制失败", str(exc))
            return
        self._set_status("飞机号 @ls0514 已复制")

    def verify_admin_code(self) -> None:
        code = self.authorization_code_var.get().strip()
        if not code:
            messagebox.showwarning("缺少验证码", "请先填写管理员提供的验证码。")
            return
        verified, message, status = verify_authorization_code(code)
        self._trial_status = status
        self._apply_authorization_state()
        if not verified:
            messagebox.showerror("验证失败", message)
            return
        self.authorization_code_var.set("")
        self._set_status(message)
        messagebox.showinfo("管理员验证", message)

    def _enforce_trial(self) -> None:
        self._trial_job = None
        was_active = self._trial_status.active
        self._trial_status = check_trial()
        if not self._trial_status.active:
            cancel_event, serial = self.operations.begin()
            self.operations.finish(cancel_event, serial)
            self._set_operation_busy(False)
        self._apply_authorization_state()
        if was_active and not self._trial_status.active:
            self._show_authorization_required()
        self._trial_job = self.after(60_000, self._enforce_trial)

    def _save_settings_values(self) -> str:
        selected_name = self.browser_provider_var.get()
        self.settings.browser_provider = next(
            (
                key
                for key, display_name in BROWSER_PROVIDER_NAMES.items()
                if display_name == selected_name
            ),
            "morelogin",
        )
        self.settings.morelogin_url = self.morelogin_var.get().strip()
        self.settings.adspower_url = self.adspower_var.get().strip()
        new_api_key = self.adspower_api_key_var.get().strip()
        if new_api_key != self.settings.adspower_api_key:
            self.settings.mark_api_key_dirty()
        self.settings.adspower_api_key = new_api_key
        self.settings.bitbrowser_url = self.bitbrowser_var.get().strip()
        self.settings.sender_name = self.sender_var.get().strip()
        if hasattr(self, "language_var"):
            self.settings.language = next(
                (
                    key
                    for key, name in LANGUAGES.items()
                    if name == self.language_var.get()
                ),
                self.settings.language,
            )
            set_language(self.settings.language)
        self.settings.save()
        self.workflow = Workflow(self.db, self.settings)
        return BROWSER_PROVIDER_NAMES[self.settings.browser_provider]

    def switch_browser_provider(self) -> None:
        try:
            provider_name = self._save_settings_values()
        except Exception as exc:
            self._set_status(f"切换失败：{exc}")
            messagebox.showerror("切换失败", str(exc))
            return
        self._set_status(f"窗口应用已切换为 {provider_name}")

    def save_settings(self) -> None:
        try:
            provider_name = self._save_settings_values()
        except Exception as exc:
            self._set_status(f"设置保存失败：{exc}")
            messagebox.showerror("设置保存失败", str(exc))
            return
        self._translate_widgets()
        self._set_status("设置已保存")

    def show_log_viewer(self) -> None:
        from ophelia_assistant.config import app_data_dir

        log_dir = app_data_dir()
        log_path = log_dir / "app.log"
        dialog = tk.Toplevel(self)
        dialog.title("运行日志与失败截图")
        dialog.geometry("780x540")
        dialog.configure(bg=PAPER)
        text = tk.Text(
            dialog,
            wrap="word",
            bg="#FFFFFF",
            fg=INK,
            relief="flat",
            padx=10,
            pady=10,
        )
        text.pack(fill="both", expand=True, padx=12, pady=(12, 6))
        try:
            lines = log_path.read_text(
                encoding="utf-8", errors="replace"
            ).splitlines()
        except OSError:
            lines = ["（暂无日志）"]
        text.insert("1.0", "\n".join(lines[-300:]))
        text.configure(state="disabled")

        def copy_log() -> None:
            content = text.get("1.0", "end-1c")
            self.clipboard_clear()
            self.clipboard_append(content)
            self._set_status("日志已复制，可粘贴给管理员")

        def open_log_folder() -> None:
            try:
                os.startfile(str(log_dir))  # type: ignore[attr-defined]
            except OSError:
                pass

        log_actions = tk.Frame(dialog, bg=PAPER)
        log_actions.pack(fill="x", padx=12, pady=(0, 8))
        ttk.Button(
            log_actions, text=tr("复制日志"), style="Soft.TButton",
            command=copy_log,
        ).pack(side="left")
        ttk.Button(
            log_actions, text=tr("打开日志文件夹"), style="Soft.TButton",
            command=open_log_folder,
        ).pack(side="left", padx=(8, 0))
        shots_dir = log_dir / "screenshots"
        shots = (
            sorted(shots_dir.glob("*.png"), reverse=True)
            if shots_dir.exists()
            else []
        )
        tk.Label(
            dialog,
            text=f"日志：{log_path}\n失败截图：{len(shots)} 张（{shots_dir}）",
            bg=PAPER,
            fg=MUTED,
            justify="left",
            font=("Microsoft YaHei UI", 8),
        ).pack(fill="x", padx=12, pady=(0, 12))

    def check_for_update(self, silent: bool = False) -> None:
        url = (self.settings.update_url or "").strip()
        if not url:
            if silent:
                return
            messagebox.showinfo(
                "检查更新",
                f"当前版本：v{__version__}\n尚未配置更新地址，请等待管理员发布更新。",
            )
            return
        if silent and self.settings.last_update_check_at:
            try:
                last_check = datetime.fromisoformat(
                    self.settings.last_update_check_at
                )
                if (
                    datetime.now(timezone.utc) - last_check
                ).total_seconds() < 86400:
                    return
            except ValueError:
                pass
        self._set_status("正在检查更新…", busy=True)

        def runner() -> None:
            try:
                response = requests.get(url, timeout=8)
                response.raise_for_status()
                payload = response.json()
            except Exception as exc:
                def fail() -> None:
                    self._set_status("检查更新失败")
                    messagebox.showerror("检查更新失败", str(exc))

                self.after(0, fail)
                return

            def finish() -> None:
                try:
                    self.settings.last_update_check_at = now_iso()
                    self.settings.save()
                except Exception:
                    pass
                remote_version = str(payload.get("version") or "").strip().lstrip("v")
                download_url = str(payload.get("url") or "")
                current_version = str(__version__).lstrip("v")
                from .update_security import verify_update_payload

                valid_manifest, manifest_reason = verify_update_payload(payload)
                if not valid_manifest:
                    self._set_status("检查更新失败")
                    messagebox.showerror(
                        "检查更新失败",
                        f"更新清单校验失败：{manifest_reason}",
                    )
                    return
                self._set_status("检查更新完成")
                if is_newer_version(remote_version, current_version):
                    if not messagebox.askyesno(
                        "发现新版本",
                        f"发现新版本：v{remote_version}\n当前版本：v{current_version}\n\n"
                        "是否下载并安装新版本？",
                        icon=messagebox.QUESTION,
                        default=messagebox.YES,
                    ):
                        return
                    if not download_url:
                        messagebox.showerror(
                            "下载失败", "更新源未提供下载地址（url）。"
                        )
                        return
                    try:
                        response = requests.get(download_url, timeout=120, stream=True)
                        response.raise_for_status()
                        if getattr(sys, "frozen", False):
                            target = (
                                Path(sys.executable).parent
                                / f"NiuMaMail-v{remote_version}.exe"
                            )
                        else:
                            target = Path(tempfile.gettempdir()) / f"NiuMaMail-v{remote_version}.exe"
                        import hashlib

                        digest = hashlib.sha256()
                        with open(target, "wb") as handle:
                            for chunk in response.iter_content(chunk_size=1 << 16):
                                handle.write(chunk)
                                digest.update(chunk)
                        expected_sha256 = str(payload.get("sha256") or "").lower()
                        if (
                            not expected_sha256
                            or digest.hexdigest().lower() != expected_sha256
                        ):
                            try:
                                target.unlink()
                            except OSError:
                                pass
                            self._set_status("下载失败：SHA-256 校验未通过")
                            messagebox.showerror(
                                "下载失败",
                                "下载文件 SHA-256 校验失败，已取消安装",
                            )
                            return
                    except Exception as exc:
                        self._set_status(f"下载失败：{exc}")
                        messagebox.showerror("下载失败", str(exc))
                        return
                    self._set_status(f"新版本已下载：{target}")
                    if messagebox.askyesno(
                        "安装新版本",
                        f"新版本已下载到：\n{target}\n\n是否立即启动新版本？",
                        default=messagebox.YES,
                    ):
                        if sys.platform == "win32":
                            os.startfile(str(target))  # type: ignore[attr-defined]
                        self.after(300, self.destroy)
                else:
                    if not silent:
                        messagebox.showinfo(
                            "检查更新", f"当前已是最新版本：v{current_version}"
                        )

            self.after(0, finish)

        threading.Thread(target=runner, daemon=True).start()

    def test_browser_connection(self) -> None:
        if self._test_busy:
            return
        self._test_busy = True
        provider_name = self._save_settings_values()
        provider = create_browser_provider(self.settings)
        self.test_connection_button.configure(text="测试中…")
        self.test_connection_button.state(["disabled"])
        self._set_status(f"正在测试 {provider_name} 本地连接", busy=True)

        def runner() -> None:
            connected = provider.ping()

            def finish() -> None:
                if not self.winfo_exists():
                    return
                self._test_busy = False
                self.test_connection_button.configure(text="测试当前连接")
                self.test_connection_button.state(["!disabled"])
                if connected:
                    if self.settings.browser_provider == "adspower":
                        self.settings.adspower_url = provider.base_url
                        self.adspower_var.set(provider.base_url)
                        self.settings.save()
                    self._set_status(f"{provider_name} 本地连接正常，已自动记住可用接口")
                else:
                    detail = getattr(provider, "last_error", "")
                    self._set_status(f"{provider_name} 本地连接失败")
                    messagebox.showerror(
                        "连接测试失败",
                        f"无法连接 {provider_name}。请确认该应用已启动且本地 API 已开启。"
                        + (f"\n\n详细信息：{detail}" if detail else ""),
                    )

            self.after(0, finish)

        threading.Thread(target=runner, daemon=True).start()

    def _on_close(self) -> None:
        choice = messagebox.askyesnocancel(
            "关闭牛马邮箱",
            "请选择关闭方式：\n\n"
            "是(Y)：后台运行（最小化到任务栏，任务继续执行）\n"
            "否(N)：直接关闭程序\n"
            "取消：继续使用",
            icon=messagebox.QUESTION,
            default=messagebox.CANCEL,
        )
        if choice is None:
            return
        if choice:
            self.iconify()
            self._set_status("已后台运行，任务继续执行；点击任务栏图标可恢复")
            return
        if getattr(self, "_operation_busy", False):
            if not messagebox.askyesno(
                "任务仍在执行",
                "批量任务仍在进行中。退出后已打开的 Gmail 窗口会保留，"
                "但未完成的任务不会继续。确定退出吗？",
                icon=messagebox.WARNING,
                default=messagebox.NO,
            ):
                return
        if self._trial_job:
            self.after_cancel(self._trial_job)
            self._trial_job = None
        if getattr(self, "_busy_watchdog_job", None):
            self.after_cancel(self._busy_watchdog_job)
            self._busy_watchdog_job = None
        cancel_event, serial = self.operations.begin()
        self.operations.finish(cancel_event, serial)
        self._worker_pool.shutdown()
        self.destroy()


def main() -> None:
    App().mainloop()
